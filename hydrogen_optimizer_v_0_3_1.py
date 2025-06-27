# e-Hydrogen Cost Optimizer by KAUST (v.0.3.1)

# This program is free software: you can redistribute it and/or modify
# v.0.3.1 Release notes
# - Corrected H2 storage costs 
# - Enhanced visualization of some graphs/charts
# - Added new default values for dropdown menus
# - Added LCA calculation for optimized LCOH scenario utilizing ecoinvent (v.3.10.1) database processes
# - Added comments for Python scructure

#Libraries: 
# Hereby we state the installation of some libaries that are required to run the code. 
# Please use the requirements.txt file to install the libraries.
#pip install customtkinter
#pip install "numpy<2"
#pip install matplotlib
#pip install openpyxl
#pip install pyomo
#pip3 install tkintermapview 
#pip install tkcalendar
#pip install pandas
#pip install timezonefinder
#pip install highspy (if using HiGHS solver)
#GLPK installation (glpsol.exe) (if using GLPK solver)
#MAC:
#install homebrew
#then add homebrew to PATH running this echo "export PATH=/opt/homebrew/bin:$PATH" >> ~/.bash_profile && source ~/.bash_profile
#brew install glpk

# Import necessary libraries
from CTkScrollableDropdown import *
from tkinter import ttk, filedialog, messagebox
from tkinter import *
from tkcalendar import DateEntry
import customtkinter
from PIL import Image
from PIL import ImageTk
from tkintermapview import TkinterMapView
import numpy as np
import pandas as pd
import threading
import os
from os import path
import math
# Implement the default Matplotlib key bindings.
from matplotlib.backend_bases import key_press_handler
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,
                                               NavigationToolbar2Tk)
from matplotlib.figure import Figure
#import matplotlib
import matplotlib.pyplot as plt
import requests
import json
from io import StringIO
from pyomo.environ import *
from pyomo.environ import value 
import pytz
from timezonefinder import TimezoneFinder
from datetime import datetime
import webbrowser
from openpyxl import load_workbook
from matplotlib.sankey import Sankey
import re  

# Overriding the environment variable to set a new Brightway2 directory
dat_path = path.abspath(path.join(path.dirname(path.realpath(__file__)), "modelInputs"))
new_path = path.join(dat_path, "Brightway25_Projects")
os.environ['BRIGHTWAY2_DIR'] = new_path
brightway_db = os.getenv('BRIGHTWAY2_DIR')
print(f'Database path: {brightway_db}')

# Make sure you set the environment variable before importing Brightway25.
# This ensures Brightway recognizes the correct directory when it initializes.
import bw2data as bd
import bw2io as bi
import bw2calc as bc
import bw2analyzer as bwa

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"
customtkinter.deactivate_automatic_dpi_awareness()

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
 
        # Variables
        self.marker_list = []

        # Configure window
        self.resizable(False, False)  # Disable window resizing
        self.title("e-Hydrogen Cost Optimizer by KAUST (v.0.3.1)")
        self.geometry('1200x700+0+0')

        # Configure grid layout (3x3)
        self.grid_columnconfigure((0,1), weight=0)
        self.grid_columnconfigure(2, weight=1)  
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure((1), weight=1)

        # Set the icon for the window (application icon)
        self.image_path = path.abspath(path.join(path.dirname(path.realpath(__file__)), "images"))
        #self.iconbitmap(bitmap=path.join(self.image_path, "hydrogen_optimizer_logo_v1.ico"))
        self.iconpath = ImageTk.PhotoImage(file=path.join(self.image_path, "hydrogen_optimizer_logo_v1.png"))
        self.wm_iconbitmap()
        self.iconphoto(False, self.iconpath)

        # Create directory path variables
        self.logo_image = customtkinter.CTkImage(Image.open(path.join(self.image_path, "hydrogen_optimizer_logo_v1.png")), size=(130, 140))
        self.dat_path = path.abspath(path.join(path.dirname(path.realpath(__file__)), "modelInputs"))
        self.output_path = path.abspath(path.join(path.dirname(path.realpath(__file__)), "modelOutputs"))

        # Create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=3, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(9, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="e-Hydrogen \n Cost Optimizer", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 0))
        self.sidebar_frame_label1 = customtkinter.CTkLabel(self.sidebar_frame, text="", image=self.logo_image)
        self.sidebar_frame_label1.grid(row=1, column=0, padx=20, pady=(10, 10))
        self.home_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "home_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "home_light.png")), size=(20, 20))
        self.home_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=2, column=0, sticky="ew")

        self.electrolysis_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "electrolysis_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "electrolysis_light.png")), size=(20, 20))
        self.frame_2_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Hydrogen production",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.electrolysis_image, anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=3, column=0, sticky="ew")
        self.cost_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "cost_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "cost_light.png")), size=(20, 20))
        self.frame_3_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="LCOH Optimization",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.cost_image, anchor="w", command=self.frame_3_button_event)
        self.frame_3_button.grid(row=4, column=0, sticky="ew")


        self.magnifier_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "magnifier_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "magnifier_light.png")), size=(20, 20))
        self.frame_4_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Explore Time Series",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.magnifier_image, anchor="w", command=self.frame_4_button_event)
        self.frame_4_button.grid(row=5, column=0, sticky="ew")
        self.pie_chart_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "pie_chart_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "pie_chart_light.png")), size=(20, 20))
        self.frame_5_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="LCOH Analysis",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.pie_chart_image, anchor="w", command=self.frame_5_button_event)
        self.frame_5_button.grid(row=6, column=0, sticky="ew")
        
        self.renewables_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "renewables_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "renewables_light.png")), size=(20, 20))
        self.frame_6_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Renewables Analysis",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.renewables_image, anchor="w", command=self.frame_6_button_event)
        self.frame_6_button.grid(row=7, column=0, sticky="ew")

        self.co2_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "co2_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "co2_light.png")), size=(20, 20))
        self.frame_7_button = customtkinter.CTkButton(self.sidebar_frame, corner_radius=0, height=40, border_spacing=10, text="Life Cycle Assessment",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.co2_image, anchor="w", command=self.frame_7_button_event)
        self.frame_7_button.grid(row=8, column=0, sticky="ew")

        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=10, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=11, column=0, padx=20, pady=(0, 10))
        #self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        #self.scaling_label.grid(row=12, column=0, padx=20, pady=(10, 0))
        #self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
        #                                                       command=self.change_scaling_event)
        #self.scaling_optionemenu.grid(row=13, column=0, padx=20, pady=(0, 20))

        # Home Frame that holds the textbox, location, and system parameters
        self.home_frame = customtkinter.CTkFrame(self, fg_color="transparent")

        # Create textbox
        self.textbox = customtkinter.CTkTextbox(self.home_frame,wrap="word",width=400,height=250)
        self.textbox.grid(row=0, column=0, padx=(20, 0), pady=(20, 0), sticky="nsew")
                
        # Create location frame
        self.location_frame = customtkinter.CTkFrame(self.home_frame, fg_color="transparent")
        self.location_frame.grid(row=1, column=0, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.location_frame.grid_columnconfigure((0,1), weight=1)
        self.location_label1 = customtkinter.CTkLabel(self.location_frame, text="Location:", font=customtkinter.CTkFont(size=14, weight="bold"), anchor="nw")
        self.location_label1.grid(row=0, column=0, padx=0, pady=(0, 0),sticky="nsew")
        self.location_entry_lat = customtkinter.CTkEntry(self.location_frame, placeholder_text="Latitude")
        self.location_entry_lat.grid(row=1,column=0,padx=10)
        self.location_entry_long = customtkinter.CTkEntry(self.location_frame, placeholder_text="Longitude")
        self.location_entry_long.grid(row=1,column=1,padx=10)
        self.location_button_1 = customtkinter.CTkButton(self.location_frame, text="Submit coordinates", command=self.submit_coordinates)
        self.location_button_1.grid(row=2, column=0, padx=30, pady=10, columnspan=2,sticky="ew")
        self.location_map_widget = TkinterMapView(self.location_frame, width=370, height=300,corner_radius=0)
        self.location_map_widget.grid(row=3, column=0, columnspan=2)

        # System parameters frame
        
        # Read default values
        self.pv_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='Photovoltaics')
        self.pv_scenarios_indexed=self.pv_default_values.copy()        #set the index
        self.pv_scenarios_indexed.set_index('Scenario',inplace=True)
        self.pv_scenarios=self.pv_scenarios_indexed.index.tolist()

        self.wt_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='WindTurbine')
        self.wt_scenarios_indexed=self.wt_default_values.copy()        #set the index
        self.wt_scenarios_indexed.set_index('Scenario',inplace=True)
        self.wt_scenarios=self.wt_scenarios_indexed.index.tolist()

        self.storage_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='BatteryStorage')
        self.storage_scenarios_indexed=self.storage_default_values.copy()        #set the index
        self.storage_scenarios_indexed.set_index('Scenario',inplace=True)
        self.storage_scenarios=self.storage_scenarios_indexed.index.tolist()

        self.water_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='LCOW')
        self.water_scenarios_indexed=self.water_default_values.copy()        #set the index
        self.water_scenarios_indexed.set_index('Scenario',inplace=True)
        self.water_scenarios=self.water_scenarios_indexed.index.tolist()

        self.electrolyser_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='Electrolyser')
        self.electrolyser_scenarios_indexed=self.electrolyser_default_values.copy()        #set the index
        self.electrolyser_scenarios_indexed.set_index('Scenario',inplace=True)
        self.electrolyser_scenarios=self.electrolyser_scenarios_indexed.index.tolist()

        self.h2_storage_default_values=pd.read_excel(path.join(self.dat_path, "default_values.xlsx"),sheet_name='HydrogenStorage')
        self.h2_storage_scenarios_indexed=self.h2_storage_default_values.copy()        #set the index
        self.h2_storage_scenarios_indexed.set_index('Scenario',inplace=True)
        self.h2_storage_scenarios=self.h2_storage_scenarios_indexed.index.tolist()

        # Create frame 
        self.param_frame = customtkinter.CTkScrollableFrame(self.home_frame)
        self.param_frame.grid(row=0, column=1, rowspan=2, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.param_frame.grid_columnconfigure((3), weight=1)
        self.param_frame.grid_rowconfigure((23), weight=1)
        self.param_label1 = customtkinter.CTkLabel(self.param_frame, text="General System Parameters:", font=customtkinter.CTkFont(size=14, weight="bold"), anchor="nw")
        self.param_label1.grid(row=0, column=0, columnspan=3, padx=10, pady=(0, 0),sticky="nsew")

        self.param_label_2 = customtkinter.CTkLabel(self.param_frame, text="Simulation Lifetime", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_2.grid(row=1, column=0, padx=10,sticky="nsew")
        self.param_entry_2 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_2.grid(row=1,column=1,padx=10)
        self.param_label_2_2 = customtkinter.CTkLabel(self.param_frame, text= 'years', font=('Arial',12), anchor="w")
        self.param_label_2_2.grid(row=1, column=2,sticky="nsew")

        self.param_label_3 = customtkinter.CTkLabel(self.param_frame, text="Hydrogen Demand", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_3.grid(row=2, column=0, padx=10,sticky="nsew")
        self.param_entry_3 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_3.grid(row=2,column=1,padx=10)
        self.param_label_3_2 = customtkinter.CTkLabel(self.param_frame, text= 'kg H'+ u'\u2082'+'/day', font=('Arial',12), anchor="w")
        self.param_label_3_2.grid(row=2, column=2,sticky="nsew")

        self.param_label_4 = customtkinter.CTkLabel(self.param_frame, text="Hydrogen Utilization Ratio", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_4.grid(row=3, column=0, padx=10,sticky="nsew")
        self.param_entry_4 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_4.grid(row=3,column=1,padx=10)
        self.param_label_4_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=('Arial',12), anchor="w")
        self.param_label_4_2.grid(row=3, column=2,sticky="nsew")

        self.param_label_5 = customtkinter.CTkLabel(self.param_frame, text="Return Rate/ WACC", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_5.grid(row=4, column=0, padx=10,sticky="nsew")
        self.param_entry_5 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_5.grid(row=4,column=1,padx=10)
        self.param_label_5_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=('Arial',12), anchor="w")
        self.param_label_5_2.grid(row=4, column=2,sticky="nsew")

        self.param_label_6 = customtkinter.CTkLabel(self.param_frame, text="Renewable energy sources:", font=customtkinter.CTkFont(size=14, weight="bold"), anchor="w")
        self.param_label_6.grid(row=5, column=0,columnspan=3, padx=10, pady=(0, 0),sticky="nsew")
        self.param_label_7 = customtkinter.CTkLabel(self.param_frame, text="Photovoltaics (PV):", font=customtkinter.CTkFont(size=12), anchor="ne")
        self.param_label_7.grid(row=6, column=0, padx=10,pady=(10,0),sticky="nsew")
        self.pv_option_menu = customtkinter.CTkOptionMenu(self.param_frame, values=self.pv_scenarios, command=self.pv_scenario_set, width=190)
        self.pv_option_menu.grid(row=6, column=1, padx=20)
        CTkScrollableDropdown(self.pv_option_menu, values=self.pv_scenarios, command=self.pv_scenario_set, width=250,justify="left", button_color="transparent", frame_corner_radius=0, hover_color='darkgray')
        self.param_label_8 = customtkinter.CTkLabel(self.param_frame, text="CAPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_8.grid(row=7, column=0, padx=10,sticky="nsew")
        self.param_entry_8 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_8.grid(row=7,column=1,padx=(0))
        self.param_label_8_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_8_2.grid(row=7, column=2,sticky="nsew")
        self.param_label_8_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_9 = customtkinter.CTkLabel(self.param_frame, text="OPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_9.grid(row=8, column=0, padx=10,sticky="nsew")
        self.param_entry_9 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_9.grid(row=8,column=1,padx=(0))
        self.param_label_9_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW/year', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_9_2.grid(row=8, column=2,sticky="nsew")
        self.param_label_9_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_10 = customtkinter.CTkLabel(self.param_frame, text="Lifetime", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_10.grid(row=9, column=0, padx=10,sticky="nsew")
        self.param_entry_10 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_10.grid(row=9,column=1,padx=(0))
        self.param_label_10_2 = customtkinter.CTkLabel(self.param_frame, text= 'years', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_10_2.grid(row=9, column=2,sticky="nsew")

        self.param_label_11 = customtkinter.CTkLabel(self.param_frame, text="Wind Turbines:", font=customtkinter.CTkFont(size=12), anchor="ne")
        self.param_label_11.grid(row=10, column=0, padx=10, pady=(10, 0),sticky="nsew")
        self.wt_option_menu = customtkinter.CTkOptionMenu(self.param_frame, values=self.wt_scenarios, command=self.wt_scenario_set,width=190)
        self.wt_option_menu.grid(row=10, column=1, padx=20)
        CTkScrollableDropdown(self.wt_option_menu, values=self.wt_scenarios, command=self.wt_scenario_set, width=250,justify="left", button_color="transparent", frame_corner_radius=0, hover_color='darkgray')
        self.param_label_12 = customtkinter.CTkLabel(self.param_frame, text="CAPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_12.grid(row=11, column=0, padx=10,sticky="nsew")
        self.param_entry_12 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_12.grid(row=11,column=1,padx=(0))
        self.param_label_12_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_12_2.grid(row=11, column=2,sticky="nsew")
        self.param_label_12_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_13 = customtkinter.CTkLabel(self.param_frame, text="OPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_13.grid(row=12, column=0, padx=10,sticky="nsew")
        self.param_entry_13 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_13.grid(row=12,column=1,padx=(0))
        self.param_label_13_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW/year', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_13_2.grid(row=12, column=2,sticky="nsew")
        self.param_label_13_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_14 = customtkinter.CTkLabel(self.param_frame, text="Lifetime", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_14.grid(row=13, column=0, padx=10,sticky="nsew")
        self.param_entry_14 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_14.grid(row=13,column=1,padx=(0))
        self.param_label_14_2 = customtkinter.CTkLabel(self.param_frame, text= 'years', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_14_2.grid(row=13, column=2,sticky="nsew")
        
        self.param_label_15 = customtkinter.CTkLabel(self.param_frame, text="Battery:", font=customtkinter.CTkFont(size=14, weight="bold"), anchor="w")
        self.param_label_15.grid(row=14, column=0, padx=10, pady=(10, 0),sticky="nsew")
        self.storage_option_menu = customtkinter.CTkOptionMenu(self.param_frame, values=self.storage_scenarios, command=self.storage_scenario_set,width=190)
        self.storage_option_menu.grid(row=14, column=1, padx=20)
        self.param_label_16 = customtkinter.CTkLabel(self.param_frame, text="Duration", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_16.grid(row=15, column=0, padx=10,sticky="nsew")
        self.param_entry_16 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_16.grid(row=15,column=1,padx=(0))
        self.param_label_16_2 = customtkinter.CTkLabel(self.param_frame, text= 'hours', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_16_2.grid(row=15, column=2,sticky="nsew")
        self.param_label_17 = customtkinter.CTkLabel(self.param_frame, text="Total CAPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_17.grid(row=16, column=0, padx=10,sticky="nsew")
        self.param_entry_17 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_17.grid(row=16,column=1,padx=(0))
        self.param_label_17_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_17_2.grid(row=16, column=2,sticky="nsew")
        self.param_label_17_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_18 = customtkinter.CTkLabel(self.param_frame, text="OPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_18.grid(row=17, column=0, padx=10,sticky="nsew")
        self.param_entry_18 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_18.grid(row=17,column=1,padx=(0))
        self.param_label_18_2 = customtkinter.CTkLabel(self.param_frame, text= '$/kW/year', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_18_2.grid(row=17, column=2,sticky="nsew")
        self.param_label_18_3 = customtkinter.CTkLabel(self.param_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.param_label_19 = customtkinter.CTkLabel(self.param_frame, text="Min. Operation % Capacity", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_19.grid(row=18, column=0, padx=10,sticky="nsew")
        self.param_entry_19 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_19.grid(row=18,column=1,padx=(0))
        self.param_label_19_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_19_2.grid(row=18, column=2,sticky="nsew")
        self.param_label_20 = customtkinter.CTkLabel(self.param_frame, text="Max. Operation % Capacity", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_20.grid(row=19, column=0, padx=10,sticky="nsew")
        self.param_entry_20 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_20.grid(row=19,column=1,padx=(0))
        self.param_label_20_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_20_2.grid(row=19, column=2,sticky="nsew")
        self.param_label_21 = customtkinter.CTkLabel(self.param_frame, text="Charging Efficiency", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_21.grid(row=20, column=0, padx=10,sticky="nsew")
        self.param_entry_21 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_21.grid(row=20,column=1,padx=(0))
        self.param_label_21_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_21_2.grid(row=20, column=2,sticky="nsew")
        self.param_label_22 = customtkinter.CTkLabel(self.param_frame, text="Discharging Efficiency", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_22.grid(row=21, column=0, padx=10,sticky="nsew")
        self.param_entry_22 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_22.grid(row=21,column=1,padx=(0))
        self.param_label_22_2 = customtkinter.CTkLabel(self.param_frame, text= '%', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_22_2.grid(row=21, column=2,sticky="nsew")
        self.param_label_23 = customtkinter.CTkLabel(self.param_frame, text="Lifetime", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_23.grid(row=22, column=0, padx=10,sticky="nsew")
        self.param_entry_23 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_23.grid(row=22,column=1,padx=(0))
        self.param_label_23_2 = customtkinter.CTkLabel(self.param_frame, text= 'years', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_23_2.grid(row=22, column=2,sticky="nsew")
        self.param_label_24 = customtkinter.CTkLabel(self.param_frame, text="Max. Power Cap.", font=customtkinter.CTkFont(size=12), anchor="e")
        self.param_label_24.grid(row=23, column=0, padx=10,sticky="nsew")
        self.param_entry_24 = customtkinter.CTkEntry(self.param_frame, placeholder_text="enter value")
        self.param_entry_24.grid(row=23,column=1,padx=(0))
        self.param_label_24_2 = customtkinter.CTkLabel(self.param_frame, text= 'kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.param_label_24_2.grid(row=23, column=2,sticky="nsew")
       

        '''
        #create optimization frame
        self.opt_frame = customtkinter.CTkFrame(self.home_frame)
        self.opt_frame.grid(row=1, column=1, padx=(20, 20), pady=(20, 00), sticky="nsew")
        self.opt_frame.grid_columnconfigure((3), weight=1)
        self.opt_frame.grid_rowconfigure(4, weight=1)
        self.opt_label1 = customtkinter.CTkLabel(self.opt_frame, text="Optimization:", font=customtkinter.CTkFont(size=14, weight="bold"), anchor="nw")
        self.opt_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky="nsew")
        self.opt_label_status = customtkinter.CTkLabel(self.opt_frame, text="", font=customtkinter.CTkFont(size=10, weight="normal"), anchor="e")
        self.opt_label_status.grid(row=0, column=2,columnspan=2, padx=10, pady=(10, 0),sticky="nsew")
        self.opt_button_1 = customtkinter.CTkButton(self.opt_frame, text="Optimize LCOH: Yearly demand", width=250, command=self.optimization_yearly)
        self.opt_button_1.grid(row=1, column=0, columnspan=2, padx=10, pady=10,sticky="ew")
        self.opt_button_2 = customtkinter.CTkButton(self.opt_frame, text="Optimize LCOH: Daily demand",fg_color="#3cb0c2",hover_color="#328d9c", command=self.optimization_daily)
        self.opt_button_2.grid(row=1, column=2, columnspan=2, padx=10, pady=10,sticky="ew")
        self.pv_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "pv_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "pv_light.png")), size=(30, 30))
        self.opt_image1 = customtkinter.CTkLabel(self.opt_frame, text="" , image="")
        self.opt_image1.grid(row=2, column=0,padx=10, pady=(10, 10),sticky='e')
        self.opt_label2 = customtkinter.CTkLabel(self.opt_frame, text="" , font=customtkinter.CTkFont(size=12, weight="normal"), anchor="w", justify="left")
        self.opt_label2.grid(row=2, column=1,padx=10, pady=(10, 10),sticky="w")
        self.windturbine_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "windturbine_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "windturbine_light.png")), size=(30, 30))
        self.opt_image2 = customtkinter.CTkLabel(self.opt_frame, text="" , image="")
        self.opt_image2.grid(row=2, column=2,padx=10, pady=(10, 10),sticky='e')
        self.opt_label3 = customtkinter.CTkLabel(self.opt_frame, text="" , font=customtkinter.CTkFont(size=12, weight="normal"), anchor="w", justify="left")
        self.opt_label3.grid(row=2, column=3,padx=10, pady=(10, 10),sticky="w")
        self.my_tree3 = ttk.Treeview(self.opt_frame,selectmode="extended", height=9)
        self.my_tree3['columns']= ("Name", "Value", "Unit")
        self.my_tree3['show']='headings'
        for col in self.my_tree3["columns"]:
            self.my_tree3.heading(col,text=col)
        self.my_tree3.column('Name',anchor='w',width=130)
        self.my_tree3.column('Value',anchor='center',width=60)
        self.my_tree3.column('Unit',anchor='w',width=40)
        #self.opt_label4 = customtkinter.CTkLabel(self.opt_frame, text="" , font=customtkinter.CTkFont(size=14, weight="normal"), anchor="w", justify="left")
        #self.opt_label4.grid(row=3, column=0,columnspan=2, padx=10, pady=(10, 10),sticky="nsew")
        '''

        # create second  main frame
        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        #create water desalination frame
        self.water_desalination_frame = customtkinter.CTkFrame(self.second_frame)
        self.water_desalination_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.water_desalination_frame.grid_columnconfigure((5), weight=1)
        self.water_desalination_frame.grid_rowconfigure((2), weight=1)
        self.water_desalination_label1 = customtkinter.CTkLabel(self.water_desalination_frame, text="Water desalination:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.water_desalination_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.water_option_menu = customtkinter.CTkOptionMenu(self.water_desalination_frame, values=self.water_scenarios, command=self.water_scenario_set,width=190)
        self.water_option_menu.grid(row=1, column=1, padx=20, pady=(0,20))
        self.water_label_1 = customtkinter.CTkLabel(self.water_desalination_frame, text="LCOW", font=customtkinter.CTkFont(size=12), anchor="e")
        self.water_label_1.grid(row=1, column=2, padx=10,pady=(0,20),sticky="nsew")
        self.water_entry = customtkinter.CTkEntry(self.water_desalination_frame, placeholder_text="enter value")
        self.water_entry.grid(row=1,column=3,padx=(10),pady=(0,20))
        self.water_label_1_2 = customtkinter.CTkLabel(self.water_desalination_frame, text= '$/m3', font=customtkinter.CTkFont(size=12))
        self.water_label_1_2.grid(row=1, column=4,pady=(0,20),sticky="nsew")
        self.water_label_1_3 = customtkinter.CTkLabel(self.water_desalination_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.water_label_2= customtkinter.CTkLabel(self.water_desalination_frame, text="Water efficiency", font=customtkinter.CTkFont(size=12), anchor="e")
        self.water_label_2.grid(row=2, column=2, padx=10,pady=(0,20),sticky="nsew")
        self.water_entry_2 = customtkinter.CTkEntry(self.water_desalination_frame, placeholder_text="enter value")
        self.water_entry_2.grid(row=2,column=3,padx=(10),pady=(0,20))
        self.water_label_2_2 = customtkinter.CTkLabel(self.water_desalination_frame, text= 'l/kg H'+ u'\u2082', font=('Arial',12))
        self.water_label_2_2.grid(row=2, column=4,pady=(0,20),sticky="nsew")
        self.water_label_2_3 = customtkinter.CTkLabel(self.water_desalination_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        
        #create electrolyser frame
        self.electrolyser_frame = customtkinter.CTkFrame(self.second_frame)
        self.electrolyser_frame.grid(row=1, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.electrolyser_frame.grid_columnconfigure((4), weight=1)
        self.electrolyser_frame.grid_rowconfigure((6), weight=1)
        self.electrolyser_label1 = customtkinter.CTkLabel(self.electrolyser_frame, text="Electrolyser:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.electrolyser_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.electrolyser_option_menu = customtkinter.CTkOptionMenu(self.electrolyser_frame, values=self.electrolyser_scenarios, command=self.electrolyser_scenario_set,width=250)
        self.electrolyser_option_menu.grid(row=1, column=0, padx=20, pady=(0,20))
        self.radio_var = customtkinter.StringVar(value='NA')
        self.radio_button_1 = customtkinter.CTkRadioButton(master=self.electrolyser_frame, text='PEM', variable=self.radio_var, value='PEM')
        self.radio_button_1.grid(row=2, column=0, pady=10, padx=20, sticky="n")
        self.radio_button_2 = customtkinter.CTkRadioButton(master=self.electrolyser_frame, text='ALK',variable=self.radio_var, value='ALK')
        self.radio_button_2.grid(row=3, column=0, pady=10, padx=20, sticky="n")
        self.electrolyser_label_1 = customtkinter.CTkLabel(self.electrolyser_frame, text="CAPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.electrolyser_label_1.grid(row=1, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.electrolyser_entry = customtkinter.CTkEntry(self.electrolyser_frame, placeholder_text="enter value")
        self.electrolyser_entry.grid(row=1,column=2,padx=(10),pady=(0,20))
        self.electrolyser_label_1_2 = customtkinter.CTkLabel(self.electrolyser_frame, text= '$/kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.electrolyser_label_1_2.grid(row=1, column=3,pady=(0,20),sticky="nsew")
        self.electrolyser_label_1_3 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.electrolyser_label_2 = customtkinter.CTkLabel(self.electrolyser_frame, text="OPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.electrolyser_label_2.grid(row=2, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.electrolyser_entry_2 = customtkinter.CTkEntry(self.electrolyser_frame, placeholder_text="enter value")
        self.electrolyser_entry_2.grid(row=2,column=2,padx=(10),pady=(0,20))
        self.electrolyser_label_2_2 = customtkinter.CTkLabel(self.electrolyser_frame, text= '$/kW', font=customtkinter.CTkFont(size=12), anchor="w")
        self.electrolyser_label_2_2.grid(row=2, column=3,pady=(0,20),sticky="nsew")
        self.electrolyser_label_2_3 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.electrolyser_label_3 = customtkinter.CTkLabel(self.electrolyser_frame, text="Efficiency", font=customtkinter.CTkFont(size=12), anchor="e")
        self.electrolyser_label_3.grid(row=3, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.electrolyser_entry_3 = customtkinter.CTkEntry(self.electrolyser_frame, placeholder_text="enter value")
        self.electrolyser_entry_3.grid(row=3,column=2,padx=(10),pady=(0,20))
        self.electrolyser_label_3_2 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'kWh/kg H'+ u'\u2082', font=('Arial',12), anchor="w")
        self.electrolyser_label_3_2.grid(row=3, column=3,pady=(0,20),sticky="nsew")
        self.electrolyser_label_3_3 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.electrolyser_label_4 = customtkinter.CTkLabel(self.electrolyser_frame, text="Stack Size", font=customtkinter.CTkFont(size=12), anchor="e")
        self.electrolyser_label_4.grid(row=4, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.electrolyser_entry_4 = customtkinter.CTkEntry(self.electrolyser_frame, placeholder_text="enter value")
        self.electrolyser_entry_4.grid(row=4,column=2,padx=(10),pady=(0,20))
        self.electrolyser_label_4_2 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'MW/stack',font=customtkinter.CTkFont(size=12), anchor="w")
        self.electrolyser_label_4_2.grid(row=4, column=3,pady=(0,20),sticky="nsew")
        self.electrolyser_label_4_3 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.electrolyser_label_5 = customtkinter.CTkLabel(self.electrolyser_frame, text="Lifetime", font=customtkinter.CTkFont(size=12), anchor="e")
        self.electrolyser_label_5.grid(row=5, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.electrolyser_entry_5 = customtkinter.CTkEntry(self.electrolyser_frame, placeholder_text="enter value")
        self.electrolyser_entry_5.grid(row=5,column=2,padx=(10),pady=(0,20))
        self.electrolyser_label_5_2 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'years',font=customtkinter.CTkFont(size=12), anchor="w")
        self.electrolyser_label_5_2.grid(row=5, column=3,pady=(0,20),sticky="nsew")
        self.electrolyser_label_5_3 = customtkinter.CTkLabel(self.electrolyser_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')

        #create hydrogen storage frame
        self.h2_storage_frame = customtkinter.CTkFrame(self.second_frame)
        self.h2_storage_frame.grid(row=2, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.h2_storage_frame.grid_columnconfigure((4), weight=1)
        self.h2_storage_frame.grid_rowconfigure((6), weight=1)
        self.h2_storage_label1 = customtkinter.CTkLabel(self.h2_storage_frame, text="Hydrogen Storage:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.h2_storage_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.h2_storage_option_menu = customtkinter.CTkOptionMenu(self.h2_storage_frame, values=self.h2_storage_scenarios, command=self.h2_storage_scenario_set,width=250)
        self.h2_storage_option_menu.grid(row=1, column=0, padx=20, pady=(0,20))
        self.h2_storage_label_1 = customtkinter.CTkLabel(self.h2_storage_frame, text="CAPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.h2_storage_label_1.grid(row=1, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.h2_storage_entry = customtkinter.CTkEntry(self.h2_storage_frame, placeholder_text="enter value")
        self.h2_storage_entry.grid(row=1,column=2,padx=(10),pady=(0,20))
        self.h2_storage_label_1_2 = customtkinter.CTkLabel(self.h2_storage_frame, text= '$/kg H'+ u'\u2082', font=('Arial',12), anchor="w")
        self.h2_storage_label_1_2.grid(row=1, column=3,pady=(0,20),sticky="nsew")
        self.h2_storage_label_1_3 = customtkinter.CTkLabel(self.h2_storage_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.h2_storage_label_2 = customtkinter.CTkLabel(self.h2_storage_frame, text="OPEX", font=customtkinter.CTkFont(size=12), anchor="e")
        self.h2_storage_label_2.grid(row=2, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.h2_storage_entry_2 = customtkinter.CTkEntry(self.h2_storage_frame, placeholder_text="enter value")
        self.h2_storage_entry_2.grid(row=2,column=2,padx=(10),pady=(0,20))
        self.h2_storage_label_2_2 = customtkinter.CTkLabel(self.h2_storage_frame, text= '$/kg H'+ u'\u2082', font=('Arial',12), anchor="w")
        self.h2_storage_label_2_2.grid(row=2, column=3,pady=(0,20),sticky="nsew")
        self.h2_storage_label_2_3 = customtkinter.CTkLabel(self.h2_storage_frame, text= 'source', font=customtkinter.CTkFont(size=12,underline=True),cursor='hand2',text_color='blue')
        self.h2_storage_label_3 = customtkinter.CTkLabel(self.h2_storage_frame, text="Maximum Capacity", font=customtkinter.CTkFont(size=12), anchor="e")
        self.h2_storage_label_3.grid(row=3, column=1, padx=10,pady=(0,20),sticky="nsew")
        self.h2_storage_entry_3 = customtkinter.CTkEntry(self.h2_storage_frame, placeholder_text="enter value")
        self.h2_storage_entry_3.grid(row=3,column=2,padx=(10),pady=(0,20))
        self.h2_storage_label_3_2 = customtkinter.CTkLabel(self.h2_storage_frame, text= 'kg H'+ u'\u2082', font=('Arial',12), anchor="w")
        self.h2_storage_label_3_2.grid(row=3, column=3,pady=(0,20),sticky="nsew")
        
        # create third  main frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        #create optimization frame
        self.optimization_frame = customtkinter.CTkFrame(self.third_frame)
        self.optimization_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.optimization_frame.grid_columnconfigure((4), weight=1)
        self.optimization_frame.grid_rowconfigure((5), weight=1)
        self.optimization_label1 = customtkinter.CTkLabel(self.optimization_frame, text="LCOH Optimization", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.optimization_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.opt_seg_button_1 = customtkinter.CTkSegmentedButton(self.optimization_frame)
        self.opt_seg_button_1.configure(values=["Guarantee a Daily Demand", "Guarantee a Yearly Demand"])
        self.opt_seg_button_1.grid(row=1, column=1, padx=30, pady=10, columnspan=2, sticky="ew")
        self.opt_button_1 = customtkinter.CTkButton(self.optimization_frame, text="Optimize!",command=self.optimization)
        self.opt_button_1.grid(row=1, column=3, columnspan=2, padx=30, pady=10,sticky="ew")
        self.opt_label_status = customtkinter.CTkLabel(self.optimization_frame, text="", font=customtkinter.CTkFont(size=10, weight="normal"))
        self.opt_label_status.grid(row=0, column=3,columnspan=2, padx=10, pady=(10, 0),sticky="nsew")
        #create results frame
        self.results_frame = customtkinter.CTkFrame(self.third_frame)
        self.results_frame.grid(row=1, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.results_frame.grid_columnconfigure((4), weight=1)
        self.results_frame.grid_rowconfigure((4), weight=1)
        self.results_label1 = customtkinter.CTkLabel(self.results_frame, text="Results", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.results_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.pv_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "pv_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "pv_light.png")), size=(30, 30))
        self.opt_image1 = customtkinter.CTkLabel(self.results_frame, text="" , image='')
        self.opt_image1.grid(row=1, column=1,padx=10, pady=(10, 10),sticky='e')
        self.opt_label2 = customtkinter.CTkLabel(self.results_frame, text="" , font=customtkinter.CTkFont(size=12, weight="normal"), anchor="w", justify="left")
        self.opt_label2.grid(row=1, column=2,padx=10, pady=(10, 10),sticky="w")
        self.windturbine_image = customtkinter.CTkImage(light_image=Image.open(path.join(self.image_path, "windturbine_dark.png")),
                                                 dark_image=Image.open(path.join(self.image_path, "windturbine_light.png")), size=(30, 30))
        self.opt_image2 = customtkinter.CTkLabel(self.results_frame, text="" , image='')
        self.opt_image2.grid(row=2, column=1,padx=10, pady=(10, 10),sticky='e')
        self.opt_label3 = customtkinter.CTkLabel(self.results_frame, text="" , font=customtkinter.CTkFont(size=12, weight="normal"), anchor="w", justify="left")
        self.opt_label3.grid(row=2, column=2,padx=10, pady=(10, 10),sticky="w")
        self.my_tree3 = ttk.Treeview(self.results_frame,selectmode="extended", height=9)
        self.my_tree3['columns']= ("Name", "Value", "Unit")
        self.my_tree3['show']='headings'
        for col in self.my_tree3["columns"]:
            self.my_tree3.heading(col,text=col)
        self.my_tree3.column('Name',anchor='w',width=120)
        self.my_tree3.column('Value',anchor='center',width=60)
        self.my_tree3.column('Unit',anchor='w',width=50)

        # create fourth main frame
        self.fourth_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        #create time graph frame
        self.time_graph_frame = customtkinter.CTkFrame(self.fourth_frame)
        self.time_graph_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.time_graph_frame.grid_columnconfigure((1,2), weight=1)
        self.time_graph_frame.grid_rowconfigure((2), weight=1)
        self.time_graph_label1 = customtkinter.CTkLabel(self.time_graph_frame, text="Results (time graph):", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.time_graph_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.graph_button_1 = customtkinter.CTkButton(self.time_graph_frame, text="Graph", command=self.create_graph_timeseries)
        self.graph_button_1.grid(row=0, column=2, padx=30, pady=10,sticky="ew")
        #self.time_day_entry = customtkinter.CTkEntry(self.time_graph_frame, placeholder_text="Day of the year")
        #self.time_day_entry.grid(row=1,column=1,padx=40)
        self.time_graph_label2 = customtkinter.CTkLabel(self.time_graph_frame, text="Choose day of year (2023):", font=customtkinter.CTkFont(size=12))
        self.time_graph_label2.grid(row=0, column=1, padx=10, pady=(10, 0),sticky='nsew')
        self.calendar_1 = DateEntry(self.time_graph_frame,selectmode = 'day',year=2023,month=1,day=1)
        self.calendar_1.grid(row=1,column=1,padx=0)
        self.time_seg_button_1 = customtkinter.CTkSegmentedButton(self.time_graph_frame)
        self.time_seg_button_1.configure(values=["Day", "Week"])
        self.time_seg_button_1.grid(row=1, column=2, padx=30, pady=10, sticky="ew")

        # create fifth main frame
        self.fifth_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        #create analysis frame
        self.analysis_frame = customtkinter.CTkFrame(self.fifth_frame)
        self.analysis_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.analysis_frame.grid_columnconfigure(1, weight=1)
        self.analysis_frame.grid_rowconfigure((1), weight=1)
        self.analysis_label1 = customtkinter.CTkLabel(self.analysis_frame, text="LCOH Distribution Analysis:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.analysis_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.analysis_button_1 = customtkinter.CTkButton(self.analysis_frame, text="Graph",command=self.create_bar_chart)
        self.analysis_button_1.grid(row=0, column=1, padx=30, pady=10,sticky="ew")

        # create sixth main frame
        self.sixth_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")

        #create renewables analysis frame
        self.re_analysis_frame = customtkinter.CTkFrame(self.sixth_frame)
        self.re_analysis_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.re_analysis_frame.grid_columnconfigure(1, weight=1)
        self.re_analysis_frame.grid_rowconfigure((1,3), weight=1)
        self.re_analysis_label1 = customtkinter.CTkLabel(self.re_analysis_frame, text="Electricity produced from PV and wind turbines:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.re_analysis_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.re_analysis_button_1 = customtkinter.CTkButton(self.re_analysis_frame, text="Graph",command=self.create_re_dist)
        self.re_analysis_button_1.grid(row=0, column=1, padx=30, pady=10,sticky="ew")

        # Create seventh main frame
        self.seventh_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.seventh_frame.grid_rowconfigure((0),weight=1)
        self.seventh_frame.grid_rowconfigure((1),weight=1,minsize=550)

        # create LCIA frame
        self.lcia_frame = customtkinter.CTkFrame(self.seventh_frame, height=150)
        self.lcia_frame.grid(row=0, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.lcia_frame.grid_columnconfigure((3), weight=1)
        self.lcia_frame.grid_rowconfigure((2), weight=1)
        self.lcia_label1 = customtkinter.CTkLabel(self.lcia_frame, text="Life Cycle Impact Assessment (LCIA):", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.lcia_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        self.lcia_label2 = customtkinter.CTkLabel(self.lcia_frame, text="Climate Change / Global Warming Potential", font=customtkinter.CTkFont(size=12))
        self.lcia_label2.grid(row=1, column=0, padx=10, pady=(0, 20),sticky='nsew')
        self.lcia_option_menu_values = ['CML v4.8 2016','ReCiPe 2016 v1.03, midpoint (H)']
        self.lcia_option_menu = customtkinter.CTkOptionMenu(self.lcia_frame, values=self.lcia_option_menu_values,width=250) #add command
        self.lcia_option_menu.grid(row=1, column=1, padx=20, pady=(0,20))
        self.lcia_button_1 = customtkinter.CTkButton(self.lcia_frame, text='Perform LCA', command= self.perform_lca) #add command
        self.lcia_button_1.grid(row=1, column=2, padx=20, columnspan=2, pady=(0,20) ,sticky="ew")

        # Create LCA Results frame
        self.lca_results_frame = customtkinter.CTkFrame(self.seventh_frame,height=500)
        self.lca_results_frame.grid(row=1, column=0, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.lca_results_frame.grid_columnconfigure((2), weight=1)
        self.lca_results_frame.grid_rowconfigure((2), weight=1)
        self.lca_results_label1 = customtkinter.CTkLabel(self.lca_results_frame, text="LCA Contribution Analysis:", font=customtkinter.CTkFont(size=14, weight="bold"),anchor='nw')
        self.lca_results_label1.grid(row=0, column=0, padx=10, pady=(10, 0),sticky='nsew')
        
        # select default frame
        self.select_frame_by_name("home")
        # information section
        self.info_label1 = customtkinter.CTkLabel(self, text="King Abdullah University of Science and Technology, Clean Energy Research Platform (CERP)  ",
                                                   font=customtkinter.CTkFont(size=10, weight="normal",slant="italic"), anchor="e")
        self.info_label1.grid(row=2, column=1, columnspan=2, padx=20, pady=(10, 10),sticky="nsew")
        # set default values
        self.appearance_mode_optionemenu.set('Light')
        #self.scaling_optionemenu.set("100%")
        self.textbox.insert("0.0", "Welcome to the e-Hydrogen Cost Optimizer developed by KAUST! \n\n"
                             + "Please follow the steps: \n 1. Enter the location coordinates.\n"
                             +" 2. Type in the system parameters in the 'Home' tab.\n"
                             +" 3. Type in the water desalination, electrolyser and hydrogen storage settings in the 'Hydrogen Production' tab.\n"
                             +" 4. Optimize to find out the Levelized Cost of Hydrogen (LCOH) related to the minimum Total System Cost in the 'Optimizer' Tab, you have the option to optimize to guarantee an annual or daily hydrogen demand, choose your preference.\n"
                             +" 5. Go to the 'Explore Time Series' tab to view in detail how the electricity production happens in a day or week.\n"
                             +" 6. Go to the 'LCOH Analysis' tab to view in detail how the LCOH is distributed among the components used in the outcome model.\n"
                             +" 7. Go to the 'Renewables' tab to analyse the distribution of both Solar PV and Wind power of the location selected.\n"
                             +" 8. In order to explore further the outcome model, please go to the root folder of the app /_internal/modelOutputs/results.xlsx\n\n"
                             +" Notes:\n"
                             +" The electricity production profile from PV panels and wind turbines are extracted from year 2023 from renewables.ninja. Therefore, you need to have a connection to Internet. ")
        self.location_map_widget.set_position(22.31, 39.10) # KAUST coordinates
        self.location_map_widget.set_zoom(14)
        self.opt_seg_button_1.set("Guarantee a Daily Demand")
        self.time_seg_button_1.set("Day")
        self.coordinates_submitted= False
                

    def select_frame_by_name(self, name):
            # set button color for selected button
            self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
            self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")
            self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "frame_3" else "transparent")
            self.frame_4_button.configure(fg_color=("gray75", "gray25") if name == "frame_4" else "transparent")
            self.frame_5_button.configure(fg_color=("gray75", "gray25") if name == "frame_5" else "transparent")
            self.frame_6_button.configure(fg_color=("gray75", "gray25") if name == "frame_6" else "transparent")
            self.frame_7_button.configure(fg_color=("gray75", "gray25") if name == "frame_7" else "transparent")
            # show selected frame 
            if name == "home":
                self.home_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.home_frame.grid_columnconfigure(1, weight=1)
                self.home_frame.grid_rowconfigure((1), weight=1)
            else:
                self.home_frame.grid_forget()
            if name == "frame_2":
                self.second_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.second_frame.grid_columnconfigure(0, weight=1)
                self.second_frame.grid_rowconfigure(2, weight=1)
            else:
                self.second_frame.grid_forget()
            if name == "frame_3":
                self.third_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.third_frame.grid_columnconfigure(0, weight=1)
                self.third_frame.grid_rowconfigure(1, weight=1)
            else:
                self.third_frame.grid_forget()
            if name == "frame_4":
                self.fourth_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.fourth_frame.grid_columnconfigure(0, weight=1)
                self.fourth_frame.grid_rowconfigure(0, weight=1)
            else:
                self.fourth_frame.grid_forget()
            if name == "frame_5":
                self.fifth_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.fifth_frame.grid_columnconfigure(0, weight=1)
                self.fifth_frame.grid_rowconfigure(0, weight=1)
            else:
                self.fifth_frame.grid_forget()
            if name == "frame_6":
                self.sixth_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.sixth_frame.grid_columnconfigure(0, weight=1)
                self.sixth_frame.grid_rowconfigure(0, weight=1)
            else:
                self.sixth_frame.grid_forget()
            if name == "frame_7":
                self.seventh_frame.grid(row=0, column=1, rowspan=2, columnspan=2, padx=(0, 0), pady=(0, 0), sticky="nsew")
                self.seventh_frame.grid_columnconfigure(0, weight=1)
                self.seventh_frame.grid_rowconfigure(0, weight=1)
            else:
                self.seventh_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    def frame_4_button_event(self):
        self.select_frame_by_name("frame_4")

    def frame_5_button_event(self):
        self.select_frame_by_name("frame_5")

    def frame_6_button_event(self):
        self.select_frame_by_name("frame_6")

    def frame_7_button_event(self):
        self.select_frame_by_name("frame_7")

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def submit_coordinates(self):
        try:
            float(self.location_entry_lat.get())
            float(self.location_entry_long.get())
        except:
            messagebox.showerror("Woah!", f'At least one entry for location is not a number')
        self.coordinates_submitted = True
        self.location_map_widget.set_position(float(self.location_entry_lat.get()),float(self.location_entry_long.get()))  
        for marker in self.marker_list:
            marker.delete()
        self.current_position = self.location_map_widget.get_position()
        self.marker_list.append(self.location_map_widget.set_marker(self.current_position[0], self.current_position[1]))
        print(self.current_position)

    # Default values setting (functions)
    #PV
    def pv_scenario_set(self,pv_scenario):
        self.param_entry_8.delete(0, END)
        self.param_entry_8.insert(END,self.pv_scenarios_indexed.loc[pv_scenario,'CAPEX'])
        source_url = self.pv_scenarios_indexed.loc[pv_scenario, 'source']
        if pd.notna(source_url):  # Check if the source is not NaN or empty
            # If there is a source URL, show the label and bind the callback
            self.param_label_8_3.grid(row=7, column=3, sticky="nsew")
            # Bind the label to the open_link function to open the URL on click
            self.param_label_8_3.unbind("<Button-1>")
            self.param_label_8_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            # If there is no source URL, hide the label
            self.param_label_8_3.grid_forget()
        self.param_entry_9.delete(0, END)
        self.param_entry_9.insert(END,self.pv_scenarios_indexed.loc[pv_scenario,'OPEX'])
        source_url = self.pv_scenarios_indexed.loc[pv_scenario, 'source']
        if pd.notna(source_url): 
            self.param_label_9_3.grid(row=8, column=3, sticky="nsew")
            self.param_label_9_3.unbind("<Button-1>")
            self.param_label_9_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.param_label_9_3.grid_forget()
        self.param_entry_10.delete(0, END)
        self.param_entry_10.insert(END,self.pv_scenarios_indexed.loc[pv_scenario,'Lifetime'])
        self.pv_option_menu.set(pv_scenario)

    #WT
    def wt_scenario_set(self,wt_scenario):
        self.param_entry_12.delete(0, END)
        self.param_entry_12.insert(END,self.wt_scenarios_indexed.loc[wt_scenario,'CAPEX'])
        source_url = self.wt_scenarios_indexed.loc[wt_scenario, 'source']
        if pd.notna(source_url):  
            self.param_label_12_3.grid(row=11, column=3, sticky="nsew")
            self.param_label_12_3.unbind("<Button-1>")
            self.param_label_12_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.param_label_12_3.grid_forget()
        self.param_entry_13.delete(0, END)
        self.param_entry_13.insert(END,self.wt_scenarios_indexed.loc[wt_scenario,'OPEX'])
        source_url = self.wt_scenarios_indexed.loc[wt_scenario, 'source']
        if pd.notna(source_url): 
            self.param_label_13_3.grid(row=12, column=3, sticky="nsew")
            self.param_label_13_3.unbind("<Button-1>")
            self.param_label_13_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.param_label_13_3.grid_forget()
        self.param_entry_14.delete(0, END)
        self.param_entry_14.insert(END,self.wt_scenarios_indexed.loc[wt_scenario,'Lifetime'])
        self.wt_option_menu.set(wt_scenario)
    #Battery Storage
    def storage_scenario_set(self,storage_scenario):
        self.param_entry_16.delete(0, END)
        self.param_entry_16.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Duration'])
        self.param_entry_17.delete(0, END)
        self.param_entry_17.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Total CAPEX'])
        source_url = self.storage_scenarios_indexed.loc[storage_scenario, 'source']
        if pd.notna(source_url):  
            self.param_label_17_3.grid(row=16, column=3, sticky="nsew")
            self.param_label_17_3.unbind("<Button-1>")
            self.param_label_17_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.param_label_17_3.grid_forget()
        self.param_entry_18.delete(0, END)
        self.param_entry_18.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'OPEX'])
        source_url = self.storage_scenarios_indexed.loc[storage_scenario, 'source']
        if pd.notna(source_url):  
            self.param_label_18_3.grid(row=17, column=3, sticky="nsew")
            self.param_label_18_3.unbind("<Button-1>")
            self.param_label_18_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.param_label_18_3.grid_forget()
        self.param_entry_19.delete(0, END)
        self.param_entry_19.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Min. Operation % Capacity'])
        self.param_entry_20.delete(0, END)
        self.param_entry_20.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Max. Operation % Capacity'])
        self.param_entry_21.delete(0, END)
        self.param_entry_21.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Charging Efficiency'])
        self.param_entry_22.delete(0, END)
        self.param_entry_22.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Discharging Efficiency'])
        self.param_entry_23.delete(0, END)
        self.param_entry_23.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Lifetime'])
        self.param_entry_24.delete(0, END)
        self.param_entry_24.insert(END,self.storage_scenarios_indexed.loc[storage_scenario,'Max Power Capacity'])
    #Water
    def water_scenario_set(self,water_scenario):
        self.water_entry.delete(0, END)
        self.water_entry.insert(END,self.water_scenarios_indexed.loc[water_scenario,'LCOW'])
        source_url = self.water_scenarios_indexed.loc[water_scenario, 'source 1']
        if pd.notna(source_url):  
            self.water_label_1_3.grid(row=1, column=5, pady=(0,20), sticky="nsew")
            self.water_label_1_3.unbind("<Button-1>")
            self.water_label_1_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.water_label_1_3.grid_forget()
        self.water_entry_2.delete(0, END)
        self.water_entry_2.insert(END,self.water_scenarios_indexed.loc[water_scenario,'Water efficiency'])
        source_url = self.water_scenarios_indexed.loc[water_scenario, 'source 2']
        if pd.notna(source_url):  
            self.water_label_2_3.grid(row=2, column=5, pady=(0,20), sticky="nsew")
            self.water_label_2_3.unbind("<Button-1>")
            self.water_label_2_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.water_label_2_3.grid_forget()
    #Electrolyser
    def electrolyser_scenario_set(self,electrolyser_scenario):
        self.electrolyser_type = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'Type']
        if self.electrolyser_type == 'PEM':
            self.radio_button_2.deselect()
            self.radio_button_1.invoke()
        elif self.electrolyser_type == 'ALK':
            self.radio_button_1.deselect()
            self.radio_button_2.invoke()
        else: 
            self.radio_button_1.deselect()
            self.radio_button_2.deselect()
        self.electrolyser_entry.delete(0, END)
        self.electrolyser_entry.insert(END,self.electrolyser_scenarios_indexed.loc[electrolyser_scenario,'CAPEX'])
        source_url = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'source 1']
        if pd.notna(source_url):  
            self.electrolyser_label_1_3.grid(row=1, column=4, pady=(0,20), sticky="nsew")
            self.electrolyser_label_1_3.unbind("<Button-1>")
            self.electrolyser_label_1_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.electrolyser_label_1_3.grid_forget()
        self.electrolyser_entry_2.delete(0, END)
        self.electrolyser_entry_2.insert(END,self.electrolyser_scenarios_indexed.loc[electrolyser_scenario,'OPEX'])
        source_url = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'source 2']
        if pd.notna(source_url):  
            self.electrolyser_label_2_3.grid(row=2, column=4, pady=(0,20), sticky="nsew")
            self.electrolyser_label_2_3.unbind("<Button-1>")
            self.electrolyser_label_2_3.bind("<Button-1>",lambda e, url=source_url: self.callback(url))
        else:
            self.electrolyser_label_2_3.grid_forget()
        self.electrolyser_entry_3.delete(0, END)
        self.electrolyser_entry_3.insert(END,self.electrolyser_scenarios_indexed.loc[electrolyser_scenario,'Efficiency'])
        source_url = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'source 1']
        if pd.notna(source_url):  
            self.electrolyser_label_3_3.grid(row=3, column=4, pady=(0,20), sticky="nsew")
            self.electrolyser_label_3_3.unbind("<Button-1>")
            self.electrolyser_label_3_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.electrolyser_label_3_3.grid_forget()
        self.electrolyser_entry_4.delete(0, END)
        self.electrolyser_entry_4.insert(END,self.electrolyser_scenarios_indexed.loc[electrolyser_scenario,'Stack Size'])
        source_url = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'source 1']
        if pd.notna(source_url):  
            self.electrolyser_label_4_3.grid(row=4, column=4, pady=(0,20), sticky="nsew")
            self.electrolyser_label_4_3.unbind("<Button-1>")
            self.electrolyser_label_4_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.electrolyser_label_4_3.grid_forget()
        self.electrolyser_entry_5.delete(0, END)
        self.electrolyser_entry_5.insert(END,self.electrolyser_scenarios_indexed.loc[electrolyser_scenario,'Lifetime'])
        source_url = self.electrolyser_scenarios_indexed.loc[electrolyser_scenario, 'source 1']
        if pd.notna(source_url):  
            self.electrolyser_label_5_3.grid(row=5, column=4, pady=(0,20), sticky="nsew")
            self.electrolyser_label_5_3.unbind("<Button-1>")
            self.electrolyser_label_5_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.electrolyser_label_5_3.grid_forget()

    #Hydrogen Storage
    def h2_storage_scenario_set(self,h2_storage_scenario):
        self.h2_storage_entry.delete(0, END)
        self.h2_storage_entry.insert(END,self.h2_storage_scenarios_indexed.loc[h2_storage_scenario,'CAPEX'])
        source_url = self.h2_storage_scenarios_indexed.loc[h2_storage_scenario, 'source']
        if pd.notna(source_url):  
            self.h2_storage_label_1_3.grid(row=1, column=4, pady=(0,20), sticky="nsew")
            self.h2_storage_label_1_3.unbind("<Button-1>")
            self.h2_storage_label_1_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.h2_storage_label_1_3.grid_forget()
        self.h2_storage_entry_2.delete(0, END)
        self.h2_storage_entry_2.insert(END,self.h2_storage_scenarios_indexed.loc[h2_storage_scenario,'OPEX'])
        source_url = self.h2_storage_scenarios_indexed.loc[h2_storage_scenario, 'source']
        if pd.notna(source_url):  
            self.h2_storage_label_2_3.grid(row=2, column=4, pady=(0,20), sticky="nsew")
            self.h2_storage_label_2_3.unbind("<Button-1>")
            self.h2_storage_label_2_3.bind("<Button-1>", lambda e, url=source_url: self.callback(url))
        else:
            self.h2_storage_label_2_3.grid_forget()
        self.h2_storage_entry_3.delete(0, END)
        self.h2_storage_entry_3.insert(END,self.h2_storage_scenarios_indexed.loc[h2_storage_scenario,'Max Capacity'])

    #Define a callback function
    def callback(self,url):
        webbrowser.open_new_tab(url)

    def open_excel(self):
        #open a file
        my_file = filedialog.askopenfilename(title="Open File" 
                                            ) #if running in windows, include the filetype arg filetype=(("Excel Files", ".xlsx"),("All Files", "*.*"))
        #grab the file
        try:
             
            self.df = pd.read_excel(my_file,sheet_name='systemSettings')
            
            self.df1 = pd.read_excel(my_file,sheet_name='eyUnitSettings')
        except Exception as e:
            messagebox.showerror("Woah!", f'There was a problem! {e}')
        #print dataframes
        print(self.df.head())
        print(self.df1.head())
        #clear the treeview
        self.my_tree.delete(*self.my_tree.get_children())
        self.my_tree2.delete(*self.my_tree2.get_children())
        #show data
        df_rows = self.df.to_numpy().tolist()
        for row in df_rows:
            formatted_row = [f"{value:.2f}" if isinstance(value, (float, int)) else value for value in row]
            self.my_tree.insert("","end",values=formatted_row)
        df1_rows = self.df1.to_numpy().tolist()
        for row in df1_rows:
            self.my_tree2.insert("","end",values=row)

        self.param_tab.grid(row=1,column=0,columnspan=2,pady=(0,10),padx=5,sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky='ns')
        self.my_tree.grid(row=0,column=0,padx=20,sticky="nsew")
        self.my_tree2.grid(row=0,column=0,padx=5,sticky="ew")
        #activate button 2
        #my_button_2.configure(state=NORMAL)


        
    def update_label_status(self, text):
        self.opt_label_status.configure(text=text)

    def deleteFiles(self):
        #deleting files if test mode is activated
        if(self.TestMode):
            try:
                os.remove(path.join(self.dat_path, "data.dat"))
                os.remove(path.join(self.dat_path, "pv_data.csv"))
                os.remove(path.join(self.dat_path, "wind_data.csv"))
                os.remove(path.join(self.output_path, "results.xlsx"))
            except:
                print("Test mode activated but one of files already deleted")
    
    def retrievedata(self):
        try:
            if not self.coordinates_submitted:
                raise ValueError('Coordinates have not been submitted yet')
            print('Coordinates have been submitted. Proceeding with operations...')
        except ValueError as e:
            messagebox.showerror("Woah!", f'Coordinates have not been submitted yet')
            self.update_label_status("Optimization failed")
        #variables to retrieve data from renewables.ninja
        token = '115de2fecd16dd5808c714573e4af3a1833f83e0'
        api_base = 'https://www.renewables.ninja/api/'
        s = requests.session()
        s.headers = {'Authorization': 'Token ' + token}
        #get pv data
        url = api_base + 'data/pv'
        args = {
            'lat': self.current_position[0],
            'lon': self.current_position[1],
            'date_from': '2023-01-01',
            'date_to': '2023-12-31',
            'dataset': 'merra2',
            'capacity': 1.0,
            'system_loss': 0.1,
            'tracking': 0,
            'tilt': abs(self.current_position[0]),
            'azim': 180,
            'format': 'json',
            'local_time': 'true'
        }
        r = s.get(url, params=args)
        # Parse JSON to get a pandas.DataFrame of data and dict of metadata
        parsed_response = json.loads(r.text)
        self.pv_data = pd.read_json(StringIO(json.dumps(parsed_response['data'])), orient='index')
        metadata = parsed_response['metadata']
        self.pv_data['local_time'] = pd.to_datetime(self.pv_data['local_time']) 
        self.pv_data.set_index('local_time', inplace=True) 
        print(self.pv_data.head(24))
        print(len(self.pv_data.index))
        self.pv_data.to_csv(path.join(self.dat_path, "pv_data.csv"))
        self.pv_data_year = int(self.pv_data['electricity'].sum())
        print(self.pv_data_year)
        
        url = api_base + 'data/wind'
        args = {
            'lat': self.current_position[0],
            'lon': self.current_position[1],
            'date_from': '2023-01-01',
            'date_to': '2023-12-31',
            'capacity': 1.0,
            'height': 100,
            'turbine': 'Vestas V80 2000',
            'format': 'json',
            'local_time': 'true'
        }
        r = s.get(url, params=args)
        parsed_response = json.loads(r.text)
        self.wind_data = pd.read_json(StringIO(json.dumps(parsed_response['data'])), orient='index')
        metadata2 = parsed_response['metadata']
        self.wind_data.set_index('local_time', inplace=True) 
        print(self.wind_data.head(24))
        print(len(self.wind_data.index))
        self.wind_data.to_csv(path.join(self.dat_path, "wind_data.csv"))
        self.wind_data_year = int(self.wind_data['electricity'].sum())
        print(self.wind_data_year)
        #print total electricity produced kwh/kWp*year
        self.opt_image1.configure(image=self.pv_image)
        self.opt_label2.configure(text=f"{self.pv_data_year} kWh/kWp*year")
        self.opt_image2.configure(image=self.windturbine_image)
        self.opt_label3.configure(text=f"{self.wind_data_year} kWh/kWp*year")         

    def data_pretreatment(self):
        try:
            float(self.param_entry_2.get())
            float(self.param_entry_3.get())
            float(self.param_entry_4.get())
            float(self.param_entry_5.get())
            float(self.param_entry_8.get())
            float(self.param_entry_9.get())
            float(self.param_entry_10.get())
            float(self.param_entry_12.get())
            float(self.param_entry_13.get())
            float(self.param_entry_14.get())
            float(self.param_entry_16.get())
            float(self.param_entry_17.get())
            float(self.param_entry_18.get())
            float(self.param_entry_19.get())
            float(self.param_entry_20.get())
            float(self.param_entry_21.get())
            float(self.param_entry_22.get())
            float(self.param_entry_23.get())
            float(self.param_entry_24.get())

        except:
            messagebox.showerror("Woah!", f'At least one entry for the System parameters is not a number')
            self.update_label_status("Optimization failed")
        try:
            float(self.water_entry.get())
            float(self.water_entry_2.get())
            float(self.electrolyser_entry.get())
            float(self.electrolyser_entry_2.get())
            float(self.electrolyser_entry_3.get())
            float(self.electrolyser_entry_4.get())
            float(self.electrolyser_entry_5.get())
        except:
            messagebox.showerror("Woah!", f'At least one entry for the Electrolyser parameters is not a number')
            self.update_label_status("Optimization failed")
        
        #adding pv and wind data to a dict named inputDataset
        transmissionLosses = .02
        self.inputDataset = {}
        self.inputDataset['cfWind'] = np.array(self.wind_data["electricity"])*(1-transmissionLosses)
        self.inputDataset['cfSolar'] = np.array(self.pv_data["electricity"])*(1-transmissionLosses)
        #adding system parameters to the dict named inputDataset
        self.inputDataset['simulationLifetime'] = float(self.param_entry_2.get())
        self.inputDataset['Hydrogen_Demand'] = float(self.param_entry_3.get())
        self.inputDataset['utilizationRatio'] = float(self.param_entry_4.get())/100
        self.inputDataset['WACC'] = float(self.param_entry_5.get())/100
        self.inputDataset['capexSolar'] = float(self.param_entry_8.get())*1000
        self.inputDataset['fixedOpexSolar'] = float(self.param_entry_9.get())*1000
        self.inputDataset['reLifetime'] = float(self.param_entry_10.get())
        self.inputDataset['capexWind'] = float(self.param_entry_12.get())*1000
        self.inputDataset['fixedOpexWind'] = float(self.param_entry_13.get())*1000
        #self.inputDataset['reLifetime'] = float(self.param_entry_14.get())
        self.inputDataset['BSDuration'] = float(self.param_entry_16.get())
        self.inputDataset['capexBSpower'] = float(self.param_entry_17.get())*1000
        self.inputDataset['capexBSenergy'] = 0 
        self.inputDataset['fixedOpexBSpower'] = float(self.param_entry_18.get())*1000
        self.inputDataset['minCapacityBS'] = float(self.param_entry_19.get())/100
        self.inputDataset['maxCapacityBS'] = float(self.param_entry_20.get())/100
        self.inputDataset['bsStoreEfficiency'] = float(self.param_entry_21.get())/100
        self.inputDataset['bsDeployEfficiency'] = float(self.param_entry_22.get())/100
        self.inputDataset['bsLifetime'] = float(self.param_entry_23.get())
        self.inputDataset['bsMaxPowerCapacity'] = float(self.param_entry_24.get())/1000
        self.inputDataset['capexh2storage'] = float(self.h2_storage_entry.get())
        self.inputDataset['opexh2storage'] = float(self.h2_storage_entry_2.get())
        self.inputDataset['maxh2storage'] = float(self.h2_storage_entry_3.get())
        self.inputDataset['capexEY'] = np.zeros(1)
        self.inputDataset['capexEY'][0] = float(self.electrolyser_entry.get())*1000
        self.inputDataset['fixedOpexEY'] = np.zeros(1)
        self.inputDataset['fixedOpexEY'] [0]= float(self.electrolyser_entry_2.get())*1000
        self.inputDataset['variableOpexEY'] = np.zeros(1)
        self.inputDataset['variableOpexEY'] [0]= float(self.water_entry.get())/1000*float(self.water_entry_2.get())
        self.inputDataset['energyUseEY'] = np.zeros(1)
        self.inputDataset['energyUseEY'][0]= float(self.electrolyser_entry_3.get())/1000
        self.inputDataset['stackSize'] = np.zeros(1)
        self.inputDataset['stackSize'][0] = float(self.electrolyser_entry_4.get())
        self.inputDataset['eyLifetime']= float(self.electrolyser_entry_5.get())

        #printing the inputDataset dictionary
        for key, value in self.inputDataset.items():
            print(f"{key}: {value}")
        '''
        #adding system parameters (excel file, worksheet 1) data to inputDataset
        paramNames = self.df['Parameter Name']
        paramValues = self.df['Value']
        for paramName,paramValue in zip(paramNames,paramValues):
            self.inputDataset[paramName] = paramValue
        #adding electrolyser unit settings parameters (excel file, worksheet 2) data  to inputDataset
        pdEYDataset = self.df1
        for paramName in pdEYDataset.columns:
            if(paramName == "Name"):
                continue
            self.inputDataset[paramName] = np.zeros(len(pdEYDataset["capexEY"]))
            for index in np.arange(0,len(pdEYDataset["capexEY"])):
                self.inputDataset[paramName][index] = pdEYDataset[paramName][index]
        #printing the inputDataset dictionary
        for key, value in self.inputDataset.items():
            print(f"{key}: {value}")
        '''

    def writeDataFile(self):
        with open(path.join(self.dat_path, "data.dat"), 'w') as f:
            #horizon (time) set
            f.write('set horizon := ')
            for i in range(len(self.inputDataset["cfWind"])):
                f.write('%d ' % i)
            f.write(';\n\n')
            #ey plants set
            f.write('set eyPlants := ')
            for i in range(len(self.inputDataset["capexEY"])):
                f.write('%d ' % i)
            f.write(';\n\n')
            #simplifying writing .dat file with for loop
            paramNames = self.inputDataset.keys()
            #single param index names-for writing correct structure of .dat file
            singleParamIndexNames = ["cfWind","cfSolar","capexEY","fixedOpexEY","variableOpexEY","energyUseEY","stackSize"]
            for paramName in paramNames:
                if((paramName in ["horizon","eyPlants"])):
                    #skip names as they are sets defined above
                    continue
                elif(paramName in singleParamIndexNames):
                    #writing correct pyomo structure for re generation
                    f.write('param %s := \n' % (paramName))
                    for i in range(len(self.inputDataset[paramName])):
                        if(i != len(self.inputDataset[paramName])-1):
                            f.write('%d %f \n' % (i,self.inputDataset[paramName][i]))
                        else:
                            f.write('%d %f' % (i,self.inputDataset[paramName][i]))
                    f.write(';\n\n')
                else:
                    #all other parameters are single values
                    f.write('param %s := %f; \n' % (paramName,self.inputDataset[paramName]))
            print("Data (.dat) file completed")
    
    def pyomo_opt(self):
        #creating optimization model with pyomo abstract representation
        model = AbstractModel()

        ################### START SETS of DATA ###################
        #timesteps in simulation, this is based on capacity factor lenght from the inputDataset dictionary, 
        #using RangeSet() to define a range of numbers from pyomo. This sequence is defined by a start value, a final value, and a step size
        model.horizon = RangeSet(0,len(self.inputDataset["cfSolar"])-1)
        #defining dayly range
        model.days = RangeSet(0, len(self.inputDataset["cfSolar"]) // 24 - 1)
        #number of unique electroyzer (ey) models to select from
        model.eyPlants = RangeSet(0,len(self.inputDataset["capexEY"])-1)
        ################### END SETS of DATA ###################
        
        ################### START PARAMETERS  ###################
        #hydrogen demand on daily basis (differs form formulation which does some calculations before hand)
        model.Hydrogen_Demand = Param()
        #respective capacity factor timeseries data for wind and solar site
        #an indexed parameter can be specified by providing sets as unnamed arguments to the Param declaration
        model.cfWind = Param(model.horizon)
        model.cfSolar = Param(model.horizon)
        #energy efficiency of storing energy in battery (e.g. you need to put in 1.2 kW to store 1 kW)
        model.bsStoreEfficiency = Param()
        #energy efficiency of deploying energy from battery
        model.bsDeployEfficiency = Param()
        #CAPEX
        #Looking at CAPEX for wind and solar per MW
        model.capexWind = Param()
        model.capexSolar = Param()
        #CAPEX for electroysis depending on plant size (to capture economics of scale)
        model.capexEY = Param(model.eyPlants)
        #battery storage power CAPEX per MWh
        model.capexBSpower = Param()
        #battery storage energy CAPEX per MW
        model.capexBSenergy = Param()
        #battery duration
        model.BSDuration = Param()
        #battery maximum power capacity
        model.bsMaxPowerCapacity = Param()

        #Fixed OPEX
        #same outline above for different stages but now looking at fixed OPEX
        model.fixedOpexWind = Param()
        model.fixedOpexSolar = Param()
        model.fixedOpexEY = Param(model.eyPlants)
        model.fixedOpexBSpower = Param()
       

        #Variable OPEX
        #Now only looking at variable OPEX for EY as all the other technologies are assumed to have negligible variable OPEX
        model.variableOpexEY = Param(model.eyPlants)

        #energy consumption for each EY model type MWh/kg H2
        model.energyUseEY = Param(model.eyPlants)
        #maximum operating percentage of nameplate capacity for BS
        model.maxCapacityBS = Param()
        #minimum operating percentage of nameplate capacity for BS
        model.minCapacityBS = Param()
        #minimum operating capacity of electroyzer
        model.minCapacityEY = Param()
        #stack size (rated energy for EY-MW) from each EY model type
        model.stackSize = Param(model.eyPlants)
        #number of years the simulation will run through
        model.simulationLifetime = Param()
        #lifetime of renewable plants
        model.reLifetime = Param()
        #lifetime of electroyzers (all treated the same)
        model.eyLifetime = Param()
        #lifetime of battery storage
        model.bsLifetime = Param()
        #WACC or discount rate for plant operations
        model.WACC = Param()
        #utilization ratio of plant (not used in model but calculated for total hydrogen production)
        model.utilizationRatio = Param()

        # Define the maximum allowable power cap
        model.maxPowerCap = Param(initialize=1000)  

        #efficiency of storing hydrogen in hydrogen storage 
        model.hsDeployEfficiency = Param(initialize=1)
        #hydrogen storage capex 
        model.capexh2storage = Param()
        #hydrogen storage opex 
        model.opexh2storage = Param()
        #hydrogen max storage
        model.maxh2storage = Param()

        ################### END PARAMETERS  ###################

        ################### START DECISION VARIABLES  ###################
        #how much wind capacity to build
        model.windCapacity = Var(domain=NonNegativeReals)
        #how much solar capacity to build
        model.solarCapacity = Var(domain=NonNegativeReals)
        #how many stacks to build of model i for EY (so only integers)
        model.eyCapacity = Var(model.eyPlants, domain = NonNegativeIntegers)
        #total battery power storage capacity to build (MWh)
        model.bsPowerCapacity = Var(domain=NonNegativeReals)
        #total battery energy storage capacity to build (MW)
        model.bsEnergyCapacity = Var(domain=NonNegativeReals)
        #H2 (kg) to produce from models i at timestep t
        model.eyGen = Var(model.eyPlants,model.horizon,domain=NonNegativeReals)
        #amount of energy (MWh) to store in battery storage at timestep t
        model.bsStore = Var(model.horizon,domain=NonNegativeReals)
        #amount of energy (MWh) available in battery at timestep t
        model.bsAvail = Var(model.horizon,domain=NonNegativeReals)
        #amount of energy (MWh) to release into islanded grid at timestep t
        model.bsDeploy = Var(model.horizon,domain=NonNegativeReals)


        model.h2storageCapacity = Var(domain=NonNegativeReals)
        #amount of hydrogen (kg) to store in hydrogen storage at timestep t
        model.hsStore = Var(model.horizon,domain=NonNegativeReals)
        #amount of hydrogen (kg) available in hydrogen storage at timestep t
        model.hsAvail = Var(model.horizon,domain=NonNegativeReals)
        #amount of hydrogen (kg) to release into islanded grid at timestep t
        model.hsDeploy = Var(model.horizon,domain=NonNegativeReals)
        ################### END DECISION VARIABLES    ###################

        ###################     START OBJECTIVE FUNCTION     ###################
        #sum of the 4 components (windCosts, solarCosts, eyCosts and bsCosts) that the LCOH includes
        def windCosts(model):
            #sum of wind CAPEX multiplied by wind capacity built + operational fixed costs*number of plant years that the OPEX is applied and then factor in time value of money
            return (model.windCapacity*(sum((model.capexWind / (math.pow((1+model.WACC),j*model.reLifetime)) for j in np.arange(math.ceil(model.simulationLifetime/model.reLifetime))))
                                        + sum((model.fixedOpexWind/(math.pow((1+model.WACC),t)) for t in np.arange(model.simulationLifetime)))))

        def solarCosts(model):
            #sum of solar CAPEX multiplied by solar capacity built + operational fixed costs*number of plant years that the OPEX is applied (same process as in wind)
            return (model.solarCapacity*(sum((model.capexSolar / (math.pow((1+model.WACC),j*model.reLifetime)) for j in np.arange(math.ceil(model.simulationLifetime/model.reLifetime))))
                                        + sum((model.fixedOpexSolar/(math.pow((1+model.WACC),t)) for t in np.arange(model.simulationLifetime)))))

        def eyCosts(model):
            # have a temporary fix for running model without annual data for variable OPEX(8760/len*+(model.horizon)))-calculates the scaling factor for full year expenditures
            # stackSize[i]*eyCapacity[i] = total MW consumption of ey model i
            # need to then multiply by capexEY (in $/MW) and fixed OPEX ($/MW) including discount factor
            # for variable ($/kg H2) multiply by generation at each hour (this could be the water cost)
            return (sum(model.stackSize[i]*model.eyCapacity[i]*(sum((model.capexEY[i] / (math.pow((1+model.WACC),j*model.eyLifetime)) for j in np.arange(math.ceil(model.simulationLifetime/model.eyLifetime)))) +
                    sum((model.fixedOpexEY[i]/(math.pow((1+model.WACC),t)) for t in np.arange(model.simulationLifetime))))  for i in model.eyPlants))

        def bsCosts(model):
            return ((sum((model.bsPowerCapacity*(model.capexBSpower / (math.pow((1+model.WACC),j*model.bsLifetime)))+ model.bsEnergyCapacity*(model.capexBSenergy / (math.pow((1+model.WACC),j*model.bsLifetime)))) for j in np.arange(math.ceil(model.simulationLifetime/model.bsLifetime))) +
                    model.bsPowerCapacity*sum((model.fixedOpexBSpower/(math.pow((1+model.WACC),t)) for t in np.arange(model.simulationLifetime)))))

        def h2storageCosts(model):
            return (model.h2storageCapacity*(model.capexh2storage
                    + sum(model.opexh2storage/(math.pow((1+model.WACC),t)) for t in np.arange(model.simulationLifetime))))

        
        def minCost_rule(model):
            return (windCosts(model) + solarCosts(model) + eyCosts(model) + bsCosts(model) + h2storageCosts(model))

        model.SystemCost = Objective(rule = minCost_rule, sense = minimize)
        ###################       END OBJECTIVE FUNCTION  ###################

        ###################       START CONSTRAINTS     ###################
        #Equation 17 requires the total hydrogen produced over all hours and model types to equal the required demand of production.
        if self.daily_constraint:
            #meet the hydrogen production targets in a daily basis
            def meetHydrogenDemand(model, d):
                daily_generation = sum(model.eyGen[i, t] - model.hsStore[t] + model.hsDeploy [t]for i in model.eyPlants for t in model.horizon if (t // 24) == d)
                return daily_generation >= model.Hydrogen_Demand
            model.meetHydrogenDemandConstraint = Constraint(model.days, rule=meetHydrogenDemand)
        else:
            #meet the hydrogen production targets for demand over the entire production basis (converting total simulation run multiplied by ratio to convert to daily basis)
            def meetHydrogenDemand(model):
                return(sum((sum(model.eyGen[i,t] - model.hsStore[t] + model.hsDeploy [t] for i in model.eyPlants))for t in model.horizon)
                    == math.ceil(len(model.horizon)/24)*model.Hydrogen_Demand)
            model.meetHydrogenDemandConstraint = Constraint(rule=meetHydrogenDemand)
        '''
        # Add the constraint for maximum yearly hydrogen production
        def maxHydrogenProduction(model):
            total_production = sum(sum(model.eyGen[i, t] for i in model.eyPlants) for t in model.horizon)
            return total_production <= math.ceil(len(model.horizon)/24)*model.Hydrogen_Demand*1.1
        model.maxHydrogenProductionConstraint = Constraint(rule=maxHydrogenProduction)            
        '''

        #Equation 6
        #generate enough energy to meet all required islanded components demand
        #Note: the energy usage parameters should encapsulate the energy efficiencies of each of the stages
        def energyDemand(model,t):
            #summing up all the demand from the various energy consumption stages (MWh/kg*kg produced gives us MWh)
            return (sum(model.energyUseEY[i]*model.eyGen[i,t] for i in model.eyPlants) + (model.bsStore[t]/model.bsStoreEfficiency))

        def energyGen(model,t):
            #summing up generation from wind, solar, and battery storage (don't need to include efficiency for bs as
            # the decision variable BSdeploy is the actual quantity deployed to grid)
            return(model.cfWind[t]*model.windCapacity + model.cfSolar[t]*model.solarCapacity + model.bsDeploy[t])

        def energyRule(model,t):
            #energy generated should always be equal to or greater than energy demanded
            return(energyDemand(model,t) == energyGen(model,t))

        model.energyConstraint = Constraint(model.horizon,rule=energyRule)

        #producing hydrogen when battery is deploying energy
        def hydrogenProductionWithBattery(model,t):
            #summing up all the demand from the various energy consumption stages (MWh/kg*kg produced gives us MWh)
            return (sum(model.energyUseEY[i]*model.eyGen[i,t] for i in model.eyPlants) >=model.bsDeploy[t])
        model.hydrogenProductionWithBatteryConstraint = Constraint(model.horizon,rule=hydrogenProductionWithBattery)

        #Battery storage constraints
        # 
        # operations definition (current energy available is
        # 30% of rated energy capacity at beginning and equal to previous available amount + new energy stored - (energy deployed + energy required to deploy it))
        def bsAvailEnergyRule(model,t):
            if t == 0:
                return(model.bsAvail[0] == .3*model.bsEnergyCapacity)
            else:
                return(model.bsAvail[t] == model.bsAvail[t-1] + model.bsStore[t-1] - (model.bsDeploy[t-1])/model.bsDeployEfficiency)
        model.bsAvailEnergyDefConstraint = Constraint(model.horizon,rule=bsAvailEnergyRule)

        #max available energy you can have in storage is storage capacity multiplied by max operating percentage (usually 95%)
        def bsAvailUpperBoundRule(model,t):
            return(model.bsAvail[t] <= model.maxCapacityBS*model.bsEnergyCapacity)
        model.bsAvailUpperBoundConstraint = Constraint(model.horizon,rule=bsAvailUpperBoundRule)

        #min state of charge of battery is storage capacity multiplied by min operating percentage
        def bsAvailLowerBoundRule(model,t):
            return(model.bsAvail[t] >= model.minCapacityBS*model.bsEnergyCapacity)
        model.bsAvailLowerBoundConstraint = Constraint(model.horizon,rule=bsAvailLowerBoundRule)


        #max amount of energy you can store in bs is max capacity - current energy charge at start - energy you deploy in hour
        #for simplicity, I assume you can not store energy in the hour and deploy it within the same hour-this would require finer resolution then hourly capacity factors
        def bsStoreEnergyUpperBoundRule(model,t):
            return(model.bsStore[t] <= model.bsEnergyCapacity - model.bsAvail[t])
        model.bsStoreEnergyUpperBoundConstraint = Constraint(model.horizon,rule=bsStoreEnergyUpperBoundRule)

        #new constraint added
        #energy stored in battery should be less than the energy generated by solar and wind
        def bsStoreEnergyUpperBoundRule2(model,t):
                return(model.bsStore[t] <= model.cfWind[t]*model.windCapacity + model.cfSolar[t]*model.solarCapacity)
        model.bsStoreEnergyUpperBound2Constraint = Constraint(model.horizon,rule=bsStoreEnergyUpperBoundRule2)

        #max amount of power you can store or deploy in bs in an hour
        def bsPowerUpperBoundRule(model,t):
            return(model.bsStore[t] + model.bsDeploy[t]/model.bsDeployEfficiency <= model.bsPowerCapacity )
        model.bsPowerUpperBoundConstraint = Constraint(model.horizon,rule=bsPowerUpperBoundRule)

       # Constraint to enforce the maximum power cap
        def bsPowerUpperBoundCapRule(model, t):
            return  model.bsPowerCapacity <= model.bsMaxPowerCapacity
        model.bsPowerUpperBoundCapConstraint = Constraint(model.horizon, rule=bsPowerUpperBoundCapRule)

        #energy deployed (and required) from storage must be less than energy available
        def bsDeployUpperBoundRule(model,t):
            return(model.bsDeploy[t]/model.bsDeployEfficiency <= model.bsAvail[t])
        model.bsDeployUpperBoundConstraint = Constraint(model.horizon,rule=bsDeployUpperBoundRule)

        #maximum energy to power ratio of 10 hours
        def bsPowerEnergyRatioMaxRule(model):
            return(model.bsEnergyCapacity <= model.BSDuration*model.bsPowerCapacity)
        model.bsPowerEnergyRatioMaxConstraint = Constraint(rule=bsPowerEnergyRatioMaxRule)

        #Hydrogen storage constraints
        
        #Availability
        def hsAvailEnergyRule(model,t):
            if t == 0:
                return(model.hsAvail[0] == 0)
            else:
                return(model.hsAvail[t] == model.hsAvail[t-1] + model.hsStore[t-1] - (model.hsDeploy[t-1])/model.hsDeployEfficiency)
        model.hsAvailEnergyDefConstraint = Constraint(model.horizon,rule=hsAvailEnergyRule)
        
        #max available hydrogen you can have store is storage capacity 
        def hsAvailUpperBoundRule(model,t):
            return(model.hsAvail[t] <= model.h2storageCapacity)
        model.hsAvailUpperBoundConstraint = Constraint(model.horizon,rule=hsAvailUpperBoundRule)

        #max amount of hydrogen you can store is max capacity - current hydrogen at start - hydrogen you deploy in hour
        def hsStoreEnergyUpperBoundRule(model,t):
            return(model.hsStore[t] <= model.h2storageCapacity - model.hsAvail[t])
        model.hsStoreEnergyUpperBoundConstraint = Constraint(model.horizon,rule=hsStoreEnergyUpperBoundRule)

        #hydrogen deployed (and required) from storage must be less than hydrogen available
        def hsDeployUpperBoundRule(model,t):
            return(model.hsDeploy[t]/model.hsDeployEfficiency <= model.hsAvail[t])
        model.hsDeployUpperBoundConstraint = Constraint(model.horizon,rule=hsDeployUpperBoundRule)

        #hydrogen stored in storage should be less than the hydrogen generated by electrolyser
        def hsStoreEnergyUpperBoundRule2(model,t):
                return(model.hsStore[t] <= sum(model.eyGen[i,t] for i in model.eyPlants))
        model.hsStoreEnergyUpperBound2Constraint = Constraint(model.horizon,rule=hsStoreEnergyUpperBoundRule2)

        # hydrogen storage is less than the maximum storage capacity
        def hsUpperBoundCapRule(model, t):
            return  model.h2storageCapacity <= model.maxh2storage
        model.hsUpperBoundCapConstraint = Constraint(model.horizon, rule=hsUpperBoundCapRule)

        #hydrogen production (in kg) from all the stacks in model i must be less than total number built*output per unit (have to convert stackSize (MW) to kg (and thus divide by energyUsage (MWh/kg)))
        #includes running at 110% of capacity
        def eyGenUpperBoundRule(model,i,t):
            return(model.eyGen[i,t] <= 1.1*((model.stackSize[i]*model.eyCapacity[i])/model.energyUseEY[i]))
        model.eyGenUpperBoundConstraint = Constraint(model.eyPlants,model.horizon,rule=eyGenUpperBoundRule)
        '''
        #max hydrogen production over a year
        def maxHydrogenProd(model):
            return(sum((sum(model.eyGen[i,t] for i in model.eyPlants))for t in model.horizon) <= math.ceil(len(model.horizon)/24)*model.Hydrogen_Demand*2)
        model.maxHydrogenProdConstraint = Constraint(rule=maxHydrogenProd)
        '''
        '''
        def eySizeRule(model,i):
            return (model.stackSize[i]*model.eyCapacity[i]) <= 2000
        model.eySizeConstraint = Constraint(model.eyPlants,rule=eySizeRule)
        '''
        ###################       END   CONSTRAINTS     ###################
        
        model.pprint()
        data = DataPortal()
        self.data_file_path=path.join(self.dat_path, "data.dat")
        #data.load(self.data_file_path, model=model)
        print(self.data_file_path)
        data.load(filename=path.join(self.dat_path, "data.dat"), model=model)
        self.instance = model.create_instance(data)
        solver = SolverFactory("appsi_highs")
        result = solver.solve(self.instance)
        #result.write()
        #self.instance.display()

    def outputFile(self):
        from pyomo.environ import value
        #setting up structure in order to get out decision variables (and objective) and save in correct excel format
        singleDecisionVariables = [     "windCapacity","solarCapacity","bsPowerCapacity","bsEnergyCapacity","h2Storage",
                                        "totalSystemCost","LCOH","windCosts","solarCosts","eyCosts","bsPowerCosts","bsEnergyCosts","h2StorageCosts",
                                        "windCapexCosts","windOpexCosts","solarCapexCosts","solarOpexCosts","eyCapexCosts","eyOpexCosts",
                                        "bsPowerCapexCosts", "bsPowerOpexCosts",
                                        "h2StorageCapex","h2StorageOpex",
                                        "waterCosts", "AveragDailyH2Prod", 'lifetime', 'reLifetime', 'pv_data_year', 'wind_data_year']



        #included wind and solar generation for simplifying data analysis and timestep
        hourlyDecisionVariables = ["windGen","solarGen","bsStore","bsAvail","bsDeploy","hsStore", "hsAvail", "hsDeploy","timestep"]

        #eyDecisionVariables =  ["eyCapacity"]
        #eyHourlyDecisionVariables = ["eyGen"]

        self.singleDvDataset = pd.DataFrame(0.0, index=np.arange(1), columns=singleDecisionVariables)
        self.hourlyDvDataset = pd.DataFrame(0.0, index=np.arange(len(self.inputDataset["cfSolar"])), columns=hourlyDecisionVariables)
        self.eySingleDvDataset = pd.DataFrame(0.0, index=np.arange(len(self.inputDataset["capexEY"])), columns=np.arange(len(self.inputDataset["capexEY"])))
        eyHourlyDvDataset = pd.DataFrame(0.0, index=np.arange(len(self.inputDataset["cfSolar"])), columns=np.arange(len(self.inputDataset["capexEY"])))

        # Assigning single values to dataframe 'singleDvDataset'
        self.singleDvDataset.loc[0,"windCapacity"] = self.instance.windCapacity.value
        self.singleDvDataset.loc[0,"solarCapacity"]= self.instance.solarCapacity.value
        self.singleDvDataset.loc[0,"bsPowerCapacity"]= self.instance.bsPowerCapacity.value
        self.singleDvDataset.loc[0,"bsEnergyCapacity"]= self.instance.bsEnergyCapacity.value
        self.singleDvDataset.loc[0,'h2Storage'] = self.instance.h2storageCapacity.value
        self.singleDvDataset.loc[0,"totalSystemCost"]= value(self.instance.SystemCost)

        #still assigning single values to df however looking at LCOH and various segments contributing
        self.totalHydrogenProduction = sum(self.instance.utilizationRatio*sum(sum(self.instance.eyGen[i, t].value for t in np.arange(len(self.instance.horizon))) for i in self.instance.eyPlants)/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime))
        self.totalHydrogenProductionNotannualized=sum(sum(sum(self.instance.eyGen[i, t].value for t in np.arange(len(self.instance.horizon))) for i in self.instance.eyPlants)for t in np.arange(self.instance.simulationLifetime)) 
        print(self.totalHydrogenProduction)
        watercosts=  sum(self.instance.variableOpexEY[i] for i in self.instance.eyPlants)
        print(watercosts)
        self.singleDvDataset.loc[0,"LCOH"]= value(self.instance.SystemCost )/self.totalHydrogenProduction + watercosts

        self.singleDvDataset.loc[0,"windCosts"]= (self.instance.windCapacity.value*(sum((self.instance.capexWind / (math.pow((1+self.instance.WACC),j*self.instance.reLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.reLifetime))))
                                        + sum((self.instance.fixedOpexWind/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))))/self.totalHydrogenProduction
        self.singleDvDataset["windCapexCosts"] = (self.instance.windCapacity.value*(sum((self.instance.capexWind / (math.pow((1+self.instance.WACC),j*self.instance.reLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.reLifetime))))))/self.totalHydrogenProduction
        self.singleDvDataset["windOpexCosts"] = (self.instance.windCapacity.value*(sum((self.instance.fixedOpexWind/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))))/self.totalHydrogenProduction

        self.singleDvDataset.loc[0,"solarCosts"]= (self.instance.solarCapacity.value*(sum((self.instance.capexSolar / (math.pow((1+self.instance.WACC),j*self.instance.reLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.reLifetime))))
                                        + sum((self.instance.fixedOpexSolar/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))))/self.totalHydrogenProduction
        self.singleDvDataset["solarCapexCosts"] = (self.instance.solarCapacity.value*(sum((self.instance.capexSolar / (math.pow((1+self.instance.WACC),j*self.instance.reLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.reLifetime))))))/self.totalHydrogenProduction
        self.singleDvDataset["solarOpexCosts"] = (self.instance.solarCapacity.value*(sum((self.instance.fixedOpexSolar/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))))/self.totalHydrogenProduction

        self.singleDvDataset.loc[0,"eyCosts"]= (sum(self.instance.stackSize[i]*self.instance.eyCapacity[i].value*(sum((self.instance.capexEY[i] / (math.pow((1+self.instance.WACC),j*self.instance.eyLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.eyLifetime)))) +
                    sum((self.instance.fixedOpexEY[i]/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime))))  for i in self.instance.eyPlants))/self.totalHydrogenProduction
        self.singleDvDataset["eyCapexCosts"] = (sum(self.instance.stackSize[i]*self.instance.eyCapacity[i].value*(sum((self.instance.capexEY[i] / (math.pow((1+self.instance.WACC),j*self.instance.eyLifetime)) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.eyLifetime)))))  for i in self.instance.eyPlants))/self.totalHydrogenProduction
        self.singleDvDataset["eyOpexCosts"] = (sum(self.instance.stackSize[i]*self.instance.eyCapacity[i].value*sum((self.instance.fixedOpexEY[i]/(math.pow((1+value(self.instance.WACC)),t))) for t in np.arange(self.instance.simulationLifetime)) for i in self.instance.eyPlants))/self.totalHydrogenProduction

        self.singleDvDataset.loc[0,"bsPowerCosts"]= (sum(self.instance.bsPowerCapacity.value*(self.instance.capexBSpower / (math.pow((1+self.instance.WACC),j*self.instance.bsLifetime)))for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.bsLifetime))) +
                                            self.instance.bsPowerCapacity.value*sum((self.instance.fixedOpexBSpower/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime))))/self.totalHydrogenProduction
        self.singleDvDataset.loc[0,"bsEnergyCosts"]= (sum(self.instance.bsEnergyCapacity.value*(self.instance.capexBSenergy / (math.pow((1+self.instance.WACC),j*self.instance.bsLifetime))) for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.bsLifetime))))/self.totalHydrogenProduction
        self.singleDvDataset["bsPowerCapexCosts"] = (sum(self.instance.bsPowerCapacity.value*(self.instance.capexBSpower / (math.pow((1+self.instance.WACC),j*self.instance.bsLifetime)))for j in np.arange(math.ceil(self.instance.simulationLifetime/self.instance.bsLifetime))))/self.totalHydrogenProduction
        self.singleDvDataset["bsPowerOpexCosts"] = (self.instance.bsPowerCapacity.value*sum((self.instance.fixedOpexBSpower/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime))))/self.totalHydrogenProduction

        self.singleDvDataset.loc[0,"h2StorageCosts"]= (self.instance.h2storageCapacity.value*self.instance.capexh2storage
                                        + self.instance.h2storageCapacity.value*sum(self.instance.opexh2storage/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))/self.totalHydrogenProduction
        self.singleDvDataset["h2StorageCapex"] = (self.instance.h2storageCapacity.value*self.instance.capexh2storage)/self.totalHydrogenProduction
        self.singleDvDataset["h2StorageOpex"] = (self.instance.h2storageCapacity.value*(sum((self.instance.opexh2storage/(math.pow((1+self.instance.WACC),t)) for t in np.arange(self.instance.simulationLifetime)))))/self.totalHydrogenProduction


        self.singleDvDataset['waterCosts'] =  watercosts

        self.singleDvDataset['AveragDailyH2Prod'] =  self.totalHydrogenProductionNotannualized/(self.instance.simulationLifetime*365)

        self.singleDvDataset['lifetime'] =  self.instance.simulationLifetime*1

        self.singleDvDataset['reLifetime'] =  self.instance.reLifetime*1

        self.singleDvDataset['pv_data_year'] =  self.pv_data_year

        self.singleDvDataset['wind_data_year'] =  self.wind_data_year



        singleDvDatasetunits = ["MW", "MW", "MW", "MWh", "kg", 
                                "$", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", 
                                "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", "$/kgH2", 
                                "$/kgH2","$/kgH2",
                                "$/kgH2", "$/kgH2", 
                                "$/kgH2","kg", 'years', 'years', 'kWh/kWhp*year', 'kWh/kWhp*year']

        #assigning hourly values to dfs
        for hour in np.arange(len(self.inputDataset["cfSolar"])):
            self.hourlyDvDataset.loc[hour,"windGen"]= self.singleDvDataset["windCapacity"][0]*self.inputDataset["cfWind"][hour]
            self.hourlyDvDataset.loc[hour,"solarGen"]= self.singleDvDataset["solarCapacity"][0]*self.inputDataset["cfSolar"][hour]
            self.hourlyDvDataset.loc[hour,"bsStore"]= self.instance.bsStore[hour].value
            self.hourlyDvDataset.loc[hour,"bsAvail"]= self.instance.bsAvail[hour].value
            self.hourlyDvDataset.loc[hour,"bsDeploy"]= self.instance.bsDeploy[hour].value
            self.hourlyDvDataset.loc[hour,"hsStore"]= self.instance.hsStore[hour].value
            self.hourlyDvDataset.loc[hour,"hsAvail"]= self.instance.hsAvail[hour].value
            self.hourlyDvDataset.loc[hour,"hsDeploy"]= self.instance.hsDeploy[hour].value

            #for later data analysis
            self.hourlyDvDataset.loc[hour,"timestep"]= hour + 1

        #assigning single dvs for ey types
        self.eySingleDvDataset = pd.DataFrame(columns=np.arange(len(self.inputDataset["capexEY"])), index=range(2))
        for eyUnit in np.arange(len(self.inputDataset["capexEY"])):
            #looking at capacity and load factor
            self.eySingleDvDataset.loc[0,eyUnit]= self.instance.stackSize[eyUnit]*self.instance.eyCapacity[eyUnit].value
            if(self.instance.eyCapacity[eyUnit].value != 0):
                self.eySingleDvDataset.loc[1,eyUnit]= sum(self.instance.eyGen[eyUnit,hour].value for hour in np.arange(len(self.instance.horizon)))/(len(self.instance.horizon)*(1/self.instance.energyUseEY[eyUnit])*self.instance.stackSize[eyUnit]*self.instance.eyCapacity[eyUnit].value)
                print('Stack size is:')
                print(self.instance.stackSize[eyUnit])
                print('ey Capacity no of stacks is:')
                print(self.instance.eyCapacity[eyUnit].value)
            else:
                self.eySingleDvDataset.loc[1,eyUnit]= 0
            self.eySingleDvDataset.loc[2,eyUnit]= self.instance.energyUseEY[eyUnit]*1000
        # naming the indices
        self.eySingleDvDataset.index = ['Capacity', 'Capacity Factor', 'Electrolyser Efficiency']
        #assigning hourly dvs for each ey unit
        for eyUnit in np.arange(len(self.inputDataset["capexEY"])):
            for hour in np.arange(len(self.inputDataset["cfSolar"])):
                eyHourlyDvDataset.loc[hour,eyUnit]= self.instance.eyGen[eyUnit,hour].value


        #now saving 4 datasets to different sheets in same excel file
        excelOutputFilePath = path.join(self.output_path, "results.xlsx")
        with pd.ExcelWriter(excelOutputFilePath) as writer:
            # Transpose singleDvDataset, add units column and write on the excel file
            transposedSingleDvDataset = self.singleDvDataset.T
            transposedSingleDvDataset.columns = ['Value']
            transposedSingleDvDataset['Unit'] = singleDvDatasetunits
            transposedSingleDvDataset.index.name = 'Variable'
            transposedSingleDvDataset.to_excel(writer,sheet_name='singleValueDvs')

            self.hourlyDvDataset.to_excel(writer,sheet_name='hourlyValueDvs')

            self.eySingleDvDataset.to_excel(writer,sheet_name='singleEyValueDvs')

            eyHourlyDvDataset.to_excel(writer,sheet_name='hourlyEyValueDvs')


            self.pv_data.index = pd.to_datetime(self.pv_data.index).tz_localize(None)
            self.pv_data.to_excel(writer,sheet_name='hourlySolarCF')

            self.wind_data.index = self.wind_data.index.tz_localize(None)
            self.wind_data.to_excel(writer,sheet_name='hourlyWindCF')


        print(f"Model output results saved to {excelOutputFilePath}")
        self.update_label_status("Optimization completed successfully!")

    def show_results(self):
        self.results_data = self.singleDvDataset
        self.results_ey_data = self.eySingleDvDataset
        lcoh = round(self.results_data.loc[0, 'LCOH'],2)
        wind_capacity = abs(round(self.results_data.loc[0, 'windCapacity'],2))
        solar_capacity = abs(round(self.results_data.loc[0, 'solarCapacity'],2))
        battery_power_capacity = abs(round(self.results_data.loc[0, 'bsPowerCapacity'],2))
        battery_energy_capacity = abs(round(self.results_data.loc[0, 'bsEnergyCapacity'],2))
        h2_storage = round(self.results_data.loc[0,'h2Storage'])
        ey_capacity_1 = self.results_ey_data.loc['Capacity', 0]
        ey_capacity_factor_1 = round(self.results_ey_data.loc['Capacity Factor', 0],2)
        daily_hydrogen_produced = round(self.results_data.loc[0,'AveragDailyH2Prod'])
        #ey_capacity_2= self.results_ey_data.loc['Capacity', 1]
        #ey_capacity_factor_2 = self.results_ey_data.loc['Capacity Factor', 1]
        self.my_tree3.delete(*self.my_tree3.get_children())
        self.my_tree3.insert(parent='',index='end',iid=0,text="",values=("LCOH",f"{lcoh:.2f}","$/kgH2"))
        self.my_tree3.insert(parent='',index='end',iid=1,text="",values=("Wind Capacity",f"{wind_capacity:.2f}","MW"))
        self.my_tree3.insert(parent='',index='end',iid=2,text="",values=("Solar Capacity",f"{solar_capacity:.2f}","MW"))
        self.my_tree3.insert(parent='',index='end',iid=3,text="",values=("Batt. Power Capacity",f"{battery_power_capacity:.2f}","MW"))
        self.my_tree3.insert(parent='',index='end',iid=4,text="",values=("Batt. Energy Capacity",f"{battery_energy_capacity:.2f}","MWh"))
        self.my_tree3.insert(parent='',index='end',iid=5,text="",values=("Electrolyser 1 Cap.",f"{ey_capacity_1:.2f}","MW"))
        self.my_tree3.insert(parent='',index='end',iid=6,text="",values=("Electrolyser 1 CF",f"{ey_capacity_factor_1:.2f}",""))
        self.my_tree3.insert(parent='',index='end',iid=7,text="",values=("H2 Storage",f"{h2_storage}","kg"))
        self.my_tree3.insert(parent='',index='end',iid=8,text="",values=("Average daily H2",f"{daily_hydrogen_produced}","kg"))
        #self.my_tree3.insert(parent='',index='end',iid=7,text="",values=("Electrolyser 2 Cap.",f"{ey_capacity_2:.2f}","MW"))
        #self.my_tree3.insert(parent='',index='end',iid=8,text="",values=("Electrolyser 2 CF",f"{ey_capacity_factor_2:.2f}",""))
        self.my_tree3.grid(row=3,column=1,columnspan=2,rowspan=2,padx=10,sticky='w')

        # Make piechart
        pie_chart_labels =  ['Wind',' PV' , ' Electrolyser' , 'Battery' , 'H2 Storage' , 'Water Desalination']
        wind_costs = abs(round(self.results_data.loc[0, 'windCosts'],2))
        solar_costs = abs(round(self.results_data.loc[0, 'solarCosts'],2))
        ey_costs = abs(round(self.results_data.loc[0, 'eyCosts'],2))
        battery_power_costs = abs(round(self.results_data.loc[0, 'bsPowerCosts'],2))
        h2_storage_costs = abs(round(self.results_data.loc[0, 'h2StorageCosts'],2))
        water_costs = abs(round(self.results_data.loc[0, 'waterCosts'],2))
        pie_chart_values = [wind_costs,solar_costs,ey_costs,battery_power_costs,h2_storage_costs,water_costs]
        pie_chart_colors = ['lightskyblue', 'gold', 'limegreen', 'violet', 'darkcyan', 'blue']
        # Custom autopct function to hide percentages for zero values
        def autopct_format(pct):
            return ('%1.1f%%' % pct) if pct > 0 else ''
        fig2 = Figure(figsize=(1, 1), dpi=72)
        ax2 = fig2.add_subplot()
        ax2.pie(pie_chart_values, autopct=autopct_format, pctdistance=1.1, radius=1, textprops={'fontsize':12}, colors=pie_chart_colors)
        ax2.legend(labels=pie_chart_labels, fontsize="8", bbox_to_anchor=(0., 0.), loc='lower left',
              ncols=2, mode="expand", borderaxespad=0.)
        ax2.set_title('Distribution Analysis of LCOH')
        # Embed the figure into the Tkinter frame
        self.canvas = FigureCanvasTkAgg(fig2, self.results_frame)
        self.canvas.draw()
        # pack_toolbar=False will make it easier to use a layout manager later on.
        toolbar = NavigationToolbar2Tk(self.canvas, self.results_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas.mpl_connect("key_press_event", key_press_handler)
        self.canvas.get_tk_widget().grid(row=1, column=3, columnspan=2, rowspan=4, padx=40, pady=(10, 0), sticky='nsew')
        toolbar.grid(row=5, column=3, columnspan=2,padx=40,sticky='ew')
        
    def create_graph_timeseries(self):
        self.get_utc_offset()
        self.graph_data = pd.read_excel(path.join(self.output_path, "results.xlsx"),sheet_name='hourlyValueDvs')
        self.graph_data2= pd.read_excel(path.join(self.output_path, "results.xlsx"),sheet_name='hourlyEyValueDvs')
        # Create a DatetimeIndex starting from "2023-01-01 00:00:00" with hourly frequency
        start_time = pd.Timestamp("2023-01-01 00:00:00")
        # Create a datetime range for 8760 hours (1 year of hourly data)
        date_range = pd.date_range(start=start_time, periods=len(self.graph_data), freq='H')
        # Adjust the date range by the UTC offset
        date_range_local = date_range + pd.DateOffset(hours=self.utc_offset)
        # Replace the first column of self.graph_data with the adjusted date range
        self.graph_data.insert(0, 'DateTime', date_range_local)
        self.graph_data2.insert(0,'DateTime', date_range_local)
        print (self.graph_data.head())
        selected_date = self.calendar_1.get_date()  # Get the selected date from DateEntry
        day_of_year = selected_date.timetuple().tm_yday  # Get the day number from timetuple
        print(f"Selected Date: {selected_date}, Day Number: {day_of_year}")
        if self.time_seg_button_1.get()=='Day':
            start_index_utc = (day_of_year - 1) * 24 #day of the year to plot
            end_index_utc = start_index_utc + 24
            start_index_local = start_index_utc - self.utc_offset
            end_index_local = end_index_utc - self.utc_offset
            # Handle wraparound for negative or too-large indices (e.g., if the offset causes it to go out of bounds)
            if start_index_local < 0:
                start_index_local = 0
            elif start_index_local>= len(self.graph_data):
                messagebox.showerror("Woah!", f'An unexpected day was requested, default value = 365')
                start_index_local = len(self.graph_data)-24
                end_index_local = end_index_local+24
                print(end_index_local)
        elif self.time_seg_button_1.get()=='Week':
            start_index_utc = (day_of_year - 1) * 24 #week of the year to plot
            end_index_utc = start_index_utc + (24*7)
            start_index_local = start_index_utc - self.utc_offset
            end_index_local = end_index_utc - self.utc_offset
            # Handle wraparound for negative or too-large indices (e.g., if the offset causes it to go out of bounds)
            if start_index_local < 0:
                start_index_local = 0
            elif start_index_local>= len(self.graph_data)-(24*7):
                messagebox.showerror("Woah!", f'An unexpected day was requested, default value = 365')
                start_index_local = len(self.graph_data)-(24*7)
                end_index_local = end_index_local+(24*7)
                print(end_index_local)
        windGen_day = self.graph_data.loc[start_index_local:end_index_local - 1, "windGen"]
        solarGen_day = self.graph_data.loc[start_index_local:end_index_local - 1, "solarGen"]
        bsAvail_day = self.graph_data.loc[start_index_local:end_index_local-1,"bsAvail"]
        bsDeploy_day = self.graph_data.loc[start_index_local:end_index_local-1,"bsDeploy"]
        hsAvail_day = self.graph_data.loc[start_index_local:end_index_local-1,"hsAvail"]
        hsDeploy_day = self.graph_data.loc[start_index_local:end_index_local-1,"hsDeploy"]
        hydrogenProd0_day = self.graph_data2.loc[start_index_local:end_index_local-1,0]
        #hydrogenProd1_day = self.graph_data2.loc[start_index_local:end_index_local-1,1]
        print(len(solarGen_day))
        time_labels = self.graph_data.loc[start_index_local:end_index_local - 1, "DateTime"]
        # Create the figure and plot
        fig1 = Figure( dpi=100)
        ax1 = fig1.add_subplot()
        #fig1,ax1=plt.subplots()
        # Plot windGen and solarGen for the full day (24 hours)
        ax1.plot(time_labels, windGen_day, marker='o', linestyle='solid', label="Wind Generation", color="blue")
        ax1.plot(time_labels, solarGen_day, marker='x', linestyle='solid', label="Solar Generation", color="orange")
        ax1.plot(time_labels, bsAvail_day, marker='^', linestyle='solid', label="Energy Available in Battery", color="darkkhaki")
        ax1.plot(time_labels, bsDeploy_day, marker='v', linestyle='solid', label="Energy Deployed from Battery", color="darkgoldenrod")
        #Plot hydrogenProduction for both electrolysers' models
        ax2 = ax1.twinx()  # instantiate a second Axes that shares the same x-axis
        ax2.plot(time_labels, hydrogenProd0_day, marker='.', linestyle='solid', label="Hydrogen Produced (1)", color="darkturquoise")
        ax2.plot(time_labels, hsAvail_day, marker='.', linestyle='solid', label="Hydrogen Stored", color="darkcyan")
        ax2.plot(time_labels, hsDeploy_day, marker='.', linestyle='solid', label="Hydrogen Deployed from Storage", color="cornflowerblue")
        #ax2.plot(time_labels, hydrogenProd1_day, marker='.', linestyle='solid', label="Hydrogen Produced (2)", color="darkcyan")
        # Set the title and labels
        ax1.set_title(r"Electricity Generation and H$_2$ production starting from day N$^{th}$ ",fontsize=12)#{n}
        ax1.set_xlabel("Hour",fontsize=10)
        ax1.set_ylabel("Generation (MWh)",fontsize=10)
        ax2.set_ylabel("Production (kg)",fontsize=10)
        # Format the x-axis to display hours in HH:MM format
        ax1.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%H:%M'))
        ax1.tick_params(axis='both', which='major', labelsize=6)  # Smaller tick labels
        # Rotate the x-axis labels for better readability
        plt.setp(ax1.get_xticklabels(), rotation=45, ha="right")
        # Add a legend to differentiate between wind and solar
        ax1.legend(loc='upper left',fontsize=6)
        ax2.legend(loc='upper right',fontsize=6)
        # Embed the figure into the Tkinter frame
        self.canvas = FigureCanvasTkAgg(fig1, self.time_graph_frame)
        self.canvas.draw()
        # pack_toolbar=False will make it easier to use a layout manager later on.
        toolbar = NavigationToolbar2Tk(self.canvas, self.time_graph_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas.mpl_connect("key_press_event", key_press_handler)
        self.canvas.get_tk_widget().grid(row=2, column=0, columnspan=3, padx=10, pady=(10, 10), sticky='nsew')
        toolbar.grid(row=3, column=0, columnspan=3,sticky='ew')
        #self.show_results()

    def get_utc_offset(self):
        lon = float(self.location_entry_long.get())
        lat = float(self.location_entry_lat.get())
        # Get the current UTC time
        current_utc_time = datetime.utcnow()
        # Find the timezone based on latitude and longitude
        tf = TimezoneFinder()
        timezone_str = tf.timezone_at(lat=lat, lng=lon)
        if timezone_str is None:
            raise ValueError("Could not determine the timezone for the given coordinates.")
        # Get the timezone object
        timezone = pytz.timezone(timezone_str)
        # Get the UTC offset for the given date
        local_time = timezone.localize(current_utc_time)
        self.utc_offset = local_time.utcoffset().total_seconds() / 3600  # Convert seconds to hours
        print(self.utc_offset) 

    def create_bar_chart(self):
        self.bar_chart_data = pd.read_excel(path.join(self.output_path, "results.xlsx"),sheet_name='singleValueDvs', index_col=0)
        species = ("LCOH",)  # Only one species
        weight_counts = {
            "Wind CAPEX": np.array([self.bar_chart_data.loc['windCapexCosts']['Value']]),  
            "Wind OPEX": np.array([self.bar_chart_data.loc['windOpexCosts']['Value']]),
            "Solar CAPEX": np.array([self.bar_chart_data.loc['solarCapexCosts']['Value']]),  
            "Solar OPEX": np.array([self.bar_chart_data.loc['solarOpexCosts']['Value']]) ,
            "Electrolyser CAPEX": np.array([self.bar_chart_data.loc['eyCapexCosts']['Value']]),  
            "Electrolyser OPEX": np.array([self.bar_chart_data.loc['eyOpexCosts']['Value']]) ,
            "Battery Power CAPEX": np.array([self.bar_chart_data.loc['bsPowerCapexCosts']['Value']]),  
            "Battery Power OPEX": np.array([self.bar_chart_data.loc['bsPowerOpexCosts']['Value']]),
            "Battery Energy CAPEX": np.array([self.bar_chart_data.loc['bsEnergyCosts']['Value']]),
            "Hydrogen Storage CAPEX": np.array([self.bar_chart_data.loc['h2StorageCapex']['Value']]),
            "Hydrogen Storage OPEX": np.array([self.bar_chart_data.loc['h2StorageOpex']['Value']]),
            "Water Desalination": np.array ([self.bar_chart_data.loc['waterCosts']['Value']])  
        }
        # Remove zero values
        weight_counts = {k: v for k, v in weight_counts.items() if v[0] > 0}
        # Define colors for each category
        colors = {
            "Wind CAPEX": "dodgerblue",
            "Wind OPEX": "lightblue",
            "Solar CAPEX": "goldenrod",
            "Solar OPEX": "gold",
            "Electrolyser CAPEX": "green",
            "Electrolyser OPEX": "lightgreen",
            "Battery Power CAPEX": "darkviolet",
            "Battery Power OPEX": "violet",
            "Battery Energy CAPEX": "hotpink",
            "Hydrogen Storage CAPEX": "darkcyan",
            "Hydrogen Storage OPEX": "cadetblue",
            "Water Desalination": "blue"
        }
        # Create a horizontal stacked bar chart
        fig3 = Figure( dpi=100)
        ax3 = fig3.add_subplot()
        bottom = np.zeros(1)
        # Plot the weights for the selected species
        for category, weight_count in weight_counts.items():
            if weight_count is not None:
                p = ax3.barh(species, weight_count, height=0.4, label=category, left=bottom, color=colors[category])
                bottom += weight_count
                ax3.bar_label(p, label_type='center', fmt='%.2f')
        # Add labels and title
        ax3.set_xlabel('$/kg H2')
        ax3.set_title('LCOH distribution',loc='left')
        fig3.legend(title='Categories',loc='outside right upper',ncol=3, fancybox=True, shadow=True)
        self.canvas3 = FigureCanvasTkAgg(fig3, self.analysis_frame)
        self.canvas3.draw()
        fig3.tight_layout()
        toolbar = NavigationToolbar2Tk(self.canvas3, self.analysis_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas3.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas3.mpl_connect("key_press_event", key_press_handler)
        self.canvas3.get_tk_widget().grid(row=1, column=0, columnspan=2, padx=10, pady=(10, 0),sticky='nsew')
        toolbar.grid(row=2, column=0, columnspan=2,sticky='ew')

    def create_re_dist(self):
        #PV data
        self.pv_data_2=pd.read_csv(path.join(self.dat_path, "pv_data.csv"),parse_dates=['local_time'],index_col=['local_time'])
        self.pv_data_2.index=pd.to_datetime(self.pv_data_2.index, errors='coerce', utc=True)
        print(self.pv_data_2.index.dtype)
        self.pv_data_2 = self.pv_data_2[self.pv_data_2.index.year == 2023]
  
        # Resample to daily data sum and save as new dataframe
        self.pv_data_2_daily = self.pv_data_2.resample('D').sum()
        print(self.pv_data_2_daily.tail())
        self.pv_data_2_daily['month'] = self.pv_data_2_daily.index.month

        fig4 = Figure( figsize=(6,3), dpi=72)
        ax4 = fig4.add_subplot()
        #fig4.subplots_adjust(left=0.1, right=0.2, top=0.1, bottom=0)  # Adjust the subplot parameters
        pv_color = 'orange'

        # Create boxplots for each month on the same axis
        for month in range(1, 13):
            self.pv_data_2_daily[self.pv_data_2_daily['month'] == month]['electricity'].plot.box(
                ax=ax4,
                positions=[month],
                widths=0.6,
                patch_artist=True,  # Enable patch artist for filling
                boxprops=dict(facecolor=pv_color, color=pv_color),  # Box color and outline color
                whiskerprops=dict(color=pv_color),  # Whisker color
                medianprops=dict(color='black'),  # Median line color
                capprops=dict(color=pv_color),  # Cap color
                showfliers=True,
                flierprops=dict(marker='o', markersize=5, linestyle='none')  # Customize outlier appearance
            )
        # Set titles and labels
        ax4.set_title("Distribution of Daily Electricity Output per kWh Installed from PV Modules", fontsize=12)
        ax4.set_xlabel("Month", fontsize=10)
        ax4.set_ylabel("Electricity (kWh)", fontsize=10)
        ax4.set_xticks(range(1, 13))
        ax4.set_xticklabels([
            'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
        ], fontsize=8)
        # Enhance grid and layout
        ax4.grid(axis='y', linestyle='--', alpha=0.7)  # Add grid lines
        #plt.tight_layout()
        #plt.show()
        self.canvas4 = FigureCanvasTkAgg(fig4, self.re_analysis_frame)
        self.canvas4.draw()
        fig4.tight_layout(pad=1)
        toolbar = NavigationToolbar2Tk(self.canvas4, self.re_analysis_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas4.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas4.mpl_connect("key_press_event", key_press_handler)
        self.canvas4.get_tk_widget().grid(row=1, column=0, columnspan=2, padx=10, pady=(10, 0),sticky='nsew')
        toolbar.grid(row=2, column=0, padx=10, columnspan=2,sticky='ew')

        #Wind data
        self.wind_data_2=pd.read_csv(path.join(self.dat_path, "wind_data.csv"),parse_dates=['local_time'],index_col=['local_time'])
        self.wind_data_2.index=pd.to_datetime(self.wind_data_2.index, errors='coerce', utc=True)
        print(self.wind_data_2.index.dtype)
        self.wind_data_2 = self.wind_data_2[self.wind_data_2.index.year == 2023]
        # Resample to daily data sum and save as new dataframe
        self.wind_data_2_daily = self.wind_data_2.resample('D').sum()
        print(self.wind_data_2_daily.tail())
        self.wind_data_2_daily['month'] = self.wind_data_2_daily.index.month

        fig5 = Figure(figsize=(7,3),dpi=72)
        ax5 = fig5.add_subplot()
        #fig5.subplots_adjust(left=0.1, right=0.2, top=0.1, bottom=0)  # Adjust the subplot parameters
        wind_color = 'lightblue'

        # Create boxplots for each month on the same axis
        for month in range(1, 13):
            self.wind_data_2_daily[self.wind_data_2_daily['month'] == month]['electricity'].plot.box(
                ax=ax5,
                positions=[month],
                widths=0.6,
                patch_artist=True,  # Enable patch artist for filling
                boxprops=dict(facecolor=wind_color, color=wind_color),  # Box color and outline color
                whiskerprops=dict(color=wind_color),  # Whisker color
                medianprops=dict(color='black'),  # Median line color
                capprops=dict(color=wind_color),  # Cap color
                showfliers=True,
                flierprops=dict(marker='o', markersize=5, linestyle='none')  # Customize outlier appearance
            )
        # Set titles and labels
        ax5.set_title("Distribution of Daily Electricity Output per kWh Installed from Wind Turbines", fontsize=12)
        ax5.set_xlabel("Month", fontsize=10)
        ax5.set_ylabel("Electricity (kWh)", fontsize=10)
        ax5.set_xticks(range(1, 13))
        ax5.set_xticklabels([
            'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
        ], fontsize=8)
        # Enhance grid and layout
        ax5.grid(axis='y', linestyle='--', alpha=0.7)  # Add grid lines
        #plt.tight_layout()
        #plt.show()
        self.canvas5 = FigureCanvasTkAgg(fig5, self.re_analysis_frame)
        self.canvas5.draw()
        fig5.tight_layout(pad=1)
        toolbar = NavigationToolbar2Tk(self.canvas5, self.re_analysis_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas5.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas5.mpl_connect("key_press_event", key_press_handler)
        self.canvas5.get_tk_widget().grid(row=3, column=0, columnspan=2, padx=10, pady=(10, 0),sticky='nsew')
        toolbar.grid(row=4, column=0, padx=10, columnspan=2,sticky='ew')

    def perform_lca(self):
        # Read the results excel file
        self.eylca_data = pd.read_excel(path.join(self.output_path, "results.xlsx"),sheet_name='singleEyValueDvs', index_col=0)
        self.lca_data = pd.read_excel(path.join(self.output_path, "results.xlsx"),sheet_name='singleValueDvs', index_col=0)


        # Modify excel file "e_Hydrogen_LCA.xlsx" containing LCI data , see 'formula' column for each activity
        lca_workbook = load_workbook(path.join(self.dat_path, 'e_Hydrogen_LCA.xlsx'))
        lca_workbook.active = lca_workbook['plant_operation']

        lca_sheet = lca_workbook.active
        electrolyser_amount_cell=lca_sheet.cell(row=12, column=2)
        electrolyser_amount_cell.value = self.eylca_data.loc['Capacity',0]/(365*self.lca_data.loc['lifetime','Value']*self.lca_data.loc['AveragDailyH2Prod','Value'])
        hydrogen_storage_amount_cell=lca_sheet.cell(row=15, column=2)
        hydrogen_storage_amount_cell.value = self.lca_data.loc['h2Storage','Value']/(365*self.lca_data.loc['lifetime','Value']*self.lca_data.loc['AveragDailyH2Prod','Value'])
        pv_amount_cell=lca_sheet.cell(row=13, column=2)
        pv_amount_cell.value=self.eylca_data.loc['Electrolyser Efficiency',0]*(self.lca_data.loc['solarCapacity','Value']/(self.lca_data.loc['solarCapacity','Value']+self.lca_data.loc['windCapacity','Value']))
        wind_amount_cell=lca_sheet.cell(row=14, column=2)
        wind_amount_cell.value=self.eylca_data.loc['Electrolyser Efficiency',0]*(self.lca_data.loc['windCapacity','Value']/(self.lca_data.loc['solarCapacity','Value']+self.lca_data.loc['windCapacity','Value']))

        lca_workbook.active = lca_workbook['pv_electricity']
        lca_sheet = lca_workbook.active
        market_photovoltaic_ammount_cell=lca_sheet.cell(row=12, column=2)
        market_photovoltaic_ammount_cell.value = 1/(self.lca_data.loc['reLifetime','Value']*0.003*1000*self.lca_data.loc['pv_data_year','Value'])
        market_tapwater_ammount_cell=lca_sheet.cell(row=13, column=2)
        market_tapwater_ammount_cell.value =  market_photovoltaic_ammount_cell.value*21.4*20
        market_wastewater_ammount_cell=lca_sheet.cell(row=14, column=2)
        market_wastewater_ammount_cell.value = market_tapwater_ammount_cell.value/1000

        lca_workbook.active = lca_workbook['wind_electricity']
        lca_sheet = lca_workbook.active
        market_windturbine_ammount_cell=lca_sheet.cell(row=12, column=2)
        market_windturbine_ammount_cell.value = 1/(self.lca_data.loc['reLifetime','Value']*4.5*1000*self.lca_data.loc['wind_data_year','Value'])
        market_wintturbine_network_ammount_cell=lca_sheet.cell(row=13, column=2)
        market_wintturbine_network_ammount_cell.value = market_windturbine_ammount_cell.value
        market_lubricatingoil_ammount_cell=lca_sheet.cell(row=14, column=2)
        market_lubricatingoil_ammount_cell.value = 354.38*self.lca_data.loc['reLifetime','Value']/(self.lca_data.loc['reLifetime','Value']*4.5*1000*self.lca_data.loc['wind_data_year','Value'])
        market_wastemineraloil_ammount_cell=lca_sheet.cell(row=15, column=2)
        market_wastemineraloil_ammount_cell.value=market_lubricatingoil_ammount_cell.value
        market_transport_cell=lca_sheet.cell(row=16, column=2)
        market_transport_cell.value=((market_lubricatingoil_ammount_cell.value/1000)/2)*(200*2*self.lca_data.loc['reLifetime','Value']/(self.lca_data.loc['reLifetime','Value']*4.5*1000*self.lca_data.loc['wind_data_year','Value']))

        lca_workbook.save(path.join(self.dat_path, 'e_Hydrogen_LCA.xlsx'))

###

        # Display the list of projects and the directory path
        print(bd.projects)  # See projects created
        print(bd.projects.dir)  # See directory path

        # Set the current project to work on
        bd.projects.set_current('e_Hydrogen_LCA')

        # Uncomment the following lines if you need to delete previous databases
        del bd.databases['e-Hydrogen LCA']

        # Import foreground data if not already imported
        if 'e-Hydrogen LCA' in bd.databases:
            print('✅ Selected foreground data already imported.')
        else:
            fg_db = bi.ExcelImporter(path.join(self.dat_path, 'e_Hydrogen_LCA.xlsx'))
            # Match foreground data internally (to itself)
            fg_db.apply_strategies()
            fg_db.match_database(fields=["name", "unit", "reference product", "location"])
            # Match foreground data to Ecoinvent datasets and biosphere
            fg_db.match_database('ecoinvent_3_10_1_selected_datasets', fields=["name", "unit", "location", "reference product"])
            fg_db.match_database("ecoinvent-3.10-biosphere", fields=["name", "categories", "location"])
            fg_db.statistics()  # Display statistics about the matching process
            print(list(fg_db.unlinked))
            fg_db.write_database()  # Save the foreground database

        # Access the foreground database and display its details
        fg_db = bd.Database('e-Hydrogen LCA')
        print("The imported foreground database is of type {} and has a length of {}.".format(type(fg_db), len(fg_db)))
        print("List of activities in the database:")
        for activity in fg_db:
            print(f"Name: {activity['name']}, Unit: {activity['unit']}, Location: {activity['location']}, Code: {activity['code']}")

        # Display the final list of databases
        print(f"📊 Final databases: {bd.databases}")

        # Select one activity from the foreground database by its code
        activity_by_code = fg_db.get("plant_operation")
        print(activity_by_code)

        # Display all exchanges (inputs and outputs) for the selected activity
        for exc in activity_by_code.exchanges():
            print(exc)

        # Select a Life Cycle Impact Assessment (LCIA) method family and method name
        selected_method = self.lcia_option_menu.get()
        print(selected_method)
        gwp_key = [
            m for m in bd.methods if "climate change" == m[1] and selected_method == m[0]
        ].pop()

        # Display details about the selected LCIA method
        method_details = bd.methods[gwp_key]
        print(f"Method key: {gwp_key}")
        print(f"Unit: {method_details.get('unit', 'No unit found')}")

        # Run the Life Cycle Assessment (LCA)
        # Prepare the functional unit and data objects for the LCA
        my_functional_unit, data_objs, _ = bd.prepare_lca_inputs(
            {activity_by_code: 1},  # Define the functional unit (1 unit of the selected activity)
            method=gwp_key,         # Use the selected LCIA method
        )

        # Initialize the LCA object
        my_lca = bc.LCA(demand=my_functional_unit, data_objs=data_objs)

        # Perform the Life Cycle Inventory (LCI) calculation
        my_lca.lci()

        # Perform the Life Cycle Impact Assessment (LCIA) calculation
        my_lca.lcia()

        # Display the final LCA score
        print(my_lca.score)

        # Perform a basic contribution analysis
        ca=bwa.utils.recursive_calculation_to_object(
            ("e-Hydrogen LCA", "plant_operation"),  # The activity to analyze
            gwp_key,
            amount=1,  # How much of the activity? (same as in our FU)
            max_level=3,  # How many levels deep in the CA do you want to go?
            cutoff=0.00001,  # Cutoff threshold for contributions
        )
        df = pd.DataFrame(ca)
        #print(df.to_string())
        df.to_csv(path.join(self.output_path,"lca_results.csv"), index=False)

        # Load the CSV file into a DataFrame
        file_path = path.join(self.output_path,"lca_results.csv")
        df = pd.read_csv(file_path)

        # Prepare the Sankey diagram data
        flows = []
        labels = []

        # Iterate through the rows to define the flows
        for index, row in df.iterrows():
            if pd.isna(row['parent']):  # Root node
                flows.append(-row['score'])
                labels.append("\n".join(row['name'].split(", ")))  # Wrap long labels
                #orientations.append(0)  # Orientation for root
                #pathlengths.append(0.25)
            elif re.match(r"^root_[a-zA-Z]$", row['label']):  # Include 'root_a', 'root_b', etc.
                flows.append(row['score'])
                labels.append("\n".join(row['name'].split(", ")))  # Wrap long labels
                #orientations.append(1)  # Orientation for child
                #pathlengths.append(0.7)

        if len(flows) == 4:
            orientations = [0,1,0,-1]
            pathlengths = [3,0.25,1,0.25]
        elif len(flows) == 5:
            orientations = [0,1,1,0,-1]
            pathlengths = [4,1,1,0.25,0.25]
        elif len(flows) == 6:
            orientations = [0,1,1,1,0,-1]
            pathlengths = [4,1,1,1,0.25,0.25]

        # Create the Sankey diagram
        fig6 = plt.figure(figsize=(8,6), dpi=72)
        ax6 = fig6.add_subplot(1,1,1,xticks=[], yticks=[], title="LCA Contribution Analysis")
        sankey = Sankey(ax=ax6, offset=0.5, head_angle=140, gap=0.3,
                format='%.3f', unit=' kg CO$_{2}$-eq', margin=2.5)  # Use LaTeX formatting for CO₂

        # Add flows and labels to the Sankey diagram
        sankey.add(flows=flows, labels=labels, orientations=orientations, pathlengths=pathlengths)

        # Finalize and display the diagram
        sankey.finish()

        # Adjust font size of the labels
        for text in ax6.texts:
            text.set_fontsize(8)  # Set the desired font size

        self.canvas6 = FigureCanvasTkAgg(fig6, self.lca_results_frame)
        self.canvas6.draw()
        fig6.tight_layout(pad=0.5)
        toolbar = NavigationToolbar2Tk(self.canvas6, self.lca_results_frame, pack_toolbar=False)
        toolbar.update()
        self.canvas6.mpl_connect(
            "key_press_event", lambda event: print(f"you pressed {event.key}"))
        self.canvas6.mpl_connect("key_press_event", key_press_handler)
        self.canvas6.get_tk_widget().grid(row=1, column=0, columnspan=3, padx=10, pady=(10, 0), sticky='nsew')
        toolbar.grid(row=2, column=0, padx=10, columnspan=3, sticky='ew')



    '''
    def optimization_yearly(self):
        self.update_label_status("Optimization in progress...")
        self.daily_constraint = False
        #start the thread for the optimization tasks
        self.optimization_thread = threading.Thread(target=self.run_optimization)
        self.optimization_thread.start()
        # Start periodic checking
        self.check_thread_status()
    '''
    def optimization(self):
        self.update_label_status("Optimization in progress...")
        if self.opt_seg_button_1.get()=='Guarantee a Daily Demand':
            self.daily_constraint = True
        else:
            self.daily_constraint = False
        #start the thread for the optimization tasks
        self.optimization_thread = threading.Thread(target=self.run_optimization)
        self.optimization_thread.start()
        # Start periodic checking
        self.check_thread_status()

    def run_optimization(self):
        self.TestMode = True
        self.deleteFiles()
        self.retrievedata()
        self.data_pretreatment()
        self.writeDataFile()
        self.pyomo_opt()  # This is your optimization function
        self.outputFile()

    def check_thread_status(self):
        if self.optimization_thread.is_alive():
            # If the thread is still running, check again after a short delay
            self.after(100, self.check_thread_status)  # Check every 100ms
        else:
            # Once the thread is done, update the label
            self.show_results()
            pass

if __name__ == "__main__":
    app = App()
    app.mainloop()